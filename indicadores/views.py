#!/usr/bin/env python
# -*- coding: utf-8 -*-
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db.models import Q
from django.db import connection, transaction
from principal.views import dictfetchall, render_to_pdf
from django.template import loader
import json
import datetime

#IMPORTAMOS BORDES 

from openpyxl.styles import Border, Side, Alignment, Color, Fill, colors, Font

from openpyxl.cell import Cell


#librerias para pdf
from io import BytesIO
from django.template import Context
from django.template.loader import get_template
from xhtml2pdf import pisa 

#Workbook nos permite crear libros en excel

from openpyxl import Workbook

from indicadores.models import Objetivos, Departamentos, Productos, Responsables, Planes
from indicadores.models import Variables, Formulas, EjesEstrategicos, PesosResponsables



def Reportes(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        context={}
        return render(request, 'indicadores/reportes.html', context)

    

#vistas
def Catalogos(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        context={
            'objetivos_list': Objetivos.objects.all().filter(estado=1),
            'departamentos_list': Departamentos.objects.all().filter(estado=1),
            'variables_list':Variables.objects.all().filter(estado=1),
            'responsables_list':Responsables.objects.all().filter(estado=1)
        }
        return render(request, 'indicadores/catalogos.html', context)


def GridObjetivosPadre(request):
    cursor = connection.cursor()
    query = "SELECT * FROM indicadores_objetivos WHERE indicadores_objetivos.objetivo_id IS null "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def SubGridObjetivosHijos(request):
    cursor = connection.cursor()
    query = "SELECT * FROM indicadores_objetivos WHERE indicadores_objetivos.objetivo_id =%s "
    params = [request.GET['objetivo_id']]
    cursor.execute(query, params)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def IngresarObjetivo(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            codigo = request.POST.get('codigo')
            nombre = request.POST.get('nombre')
            if request.POST.get('padre') == 0:
                claveForanea_Objetivo = None
            else:
                claveForanea_Objetivo = Objetivos.objects.get(pk=request.POST.get('padre'))
            indicador = request.POST.get('indicador') 
            signo = request.POST.get('signo') 
            operacion = request.POST.get('operacion')
            # 1=ponderado, 2=promedio, 3=suma 
            ponderacion = request.POST.get('ponderacion')
            # 1=anual, 2=semestral, 3=quimestral, 4=trimestral, 5=mensual 
            periodos = request.POST.get('periodos')
            # 1=si, 0=no 
            meta = request.POST.get('meta')
            if int(meta) == 1:
                meta_inicio = request.POST.get('metaInicial')
                meta_final = request.POST.get('metaFinal')
            else:
                meta_inicio = None
                meta_final = None
            # 1=activo, 0=Inactivo
            estado = request.POST.get('estado') 
            observacion = request.POST.get('comentarios') 
                  
            
            objetivos = Objetivos.objects.filter(Q (nombre=nombre) | Q (codigo=codigo)).count()
            if objetivos == 0:
                with transaction.atomic():

                    objetivos=Objetivos(
                        codigo=codigo,
                        nombre=nombre,
                        objetivo=claveForanea_Objetivo,
                        indicador=indicador,
                        signo=signo,
                        operacion=operacion,
                        ponderacion=ponderacion,
                        periodos=periodos,
                        meta=meta,
                        meta_inicio = meta_inicio,
                        meta_fin = meta_final,
                        estado=estado,
                        observacion=observacion,
                    )
                    objetivos.save()
                messages.add_message(request, messages.SUCCESS, 'Registro Creado Existosamente !!')
                return redirect('catalogos')
            else:
                messages.add_message(request, messages.WARNING, 'El Objetivo existente no puede ser creado !!')
                return redirect('catalogos')
        else:
            return redirect('logout')



def EliminarObjetivo(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        idobjetivo = request.POST.get('idobjetivo')
        with transaction.atomic():
            Objetivos.objects.filter(pk=idobjetivo).update(
                estado =0
            )
        
        data={
            'resultado':'ok_delete',
        }
        return HttpResponse(json.dumps(data), content_type='application/json')



def ConsultarObjetivo(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            idobjetivo = request.POST.get('idobjetivo')
            # consulta un objetivo por su id
            #objetivo = Objetivos.objects.values('id', 'codigo', 'nombre','objetivo','indicador','signo','operacion','ponderacion','periodos','meta','observacion','estado').all().filter(pk=idobjetivo)
            objetivo = Objetivos.objects.values().all().filter(pk=idobjetivo)
            objetivo_lista = list(objetivo)
            data={
                'resultado':'ok_select',
                'objetivo':objetivo_lista
            }
            return HttpResponse(json.dumps(data), content_type='application/json')



def ActualizarObjetivo(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            idobjetivo = request.POST.get('idobjetivo')
            codigo = request.POST.get('codigo')
            nombre = request.POST.get('nombre')
            if int(request.POST.get('padre')) == 0:
                claveForanea_Objetivo = None
            else:
                claveForanea_Objetivo = Objetivos.objects.get(pk=request.POST.get('padre'))
            indicador = request.POST.get('indicador')
            signo = request.POST.get('signo')
            operacion = request.POST.get('operacion')
            ponderacion = request.POST.get('ponderacion')
            periodos = request.POST.get('periodos')
            meta = request.POST.get('meta')
            if int(meta) == 1:
                meta_inicio = request.POST.get('metaInicial')
                meta_final = request.POST.get('metaFinal')
            else:
                meta_inicio = None
                meta_final = None
            observacion = request.POST.get('comentarios')
            estado = request.POST.get('estado')

            with transaction.atomic():
                Objetivos.objects.filter(pk=idobjetivo).update(
                    codigo = codigo,
                    nombre = nombre,
                    objetivo = claveForanea_Objetivo,
                    indicador = indicador,
                    signo = signo,
                    operacion = operacion,
                    ponderacion = ponderacion,
                    periodos = periodos,
                    meta = meta,
                    meta_inicio = meta_inicio,
                    meta_fin = meta_final,
                    observacion = observacion,
                    estado = estado,
                )
            messages.add_message(request, messages.SUCCESS, 'Objetivo Actualizado.')
            return redirect('catalogos')


def ExcelObjetivos(request):
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        #consulta objetivos
        rows = Objetivos.objects.all().filter(estado=1)
        #ESTABLECER BORDES
        thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
        )
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Objetivos"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:M2')
        ws['B2'] = 'Listado de Objetivos'
        ws['B2'].alignment = Alignment(horizontal='center')
        #ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Codigo'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Nombre'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Objetivo'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        
        ws['F3'] = 'Indicador'
        ws.cell(row=3, column=6).border = thin_border
        ws.cell(row=3, column=6).alignment = Alignment(horizontal='center')
        f3=ws['F3']
        f3.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        
        ws['G3'] = 'Signo'
        ws.cell(row=3, column=7).border = thin_border
        ws.cell(row=3, column=7).alignment = Alignment(horizontal='center')
        g3=ws['G3']
        g3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['H3'] = 'Operacion'
        ws.cell(row=3, column=8).border = thin_border
        ws.cell(row=3, column=8).alignment = Alignment(horizontal='center')
        h3=ws['H3']
        h3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['I3'] = 'ponderacion'
        ws.cell(row=3, column=9).border = thin_border
        ws.cell(row=3, column=9).alignment = Alignment(horizontal='center')
        i3=ws['I3']
        i3.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        
        ws['J3'] = 'Periodos'
        ws.cell(row=3, column=10).border = thin_border
        ws.cell(row=3, column=10).alignment = Alignment(horizontal='center')
        j3=ws['J3']
        j3.font = Font( bold=True, color=colors.DARKBLUE, size=12)
       
        ws['K3'] = 'Meta'
        ws.cell(row=3, column=11).border = thin_border
        ws.cell(row=3, column=11).alignment = Alignment(horizontal='center')
        k3=ws['K3']
        k3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['L3'] = 'Comentarios'
        ws.cell(row=3, column=12).border = thin_border
        ws.cell(row=3, column=12).alignment = Alignment(horizontal='center')
        l3=ws['L3']
        l3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for objetivo in rows:
            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = objetivo.codigo
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='left')

            ws.cell(row=cont,column=4).value = objetivo.nombre
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='left')

                        #agregamos la data, borde, alineacion
            if objetivo.objetivo:
                ws.cell(row=cont,column=5).value = objetivo.objetivo.nombre
            else:            
                ws.cell(row=cont,column=5).value = 'TODOS'
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')

                        #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=6).value = objetivo.indicador
            ws.cell(row=cont,column=6).border = thin_border
            ws.cell(row=cont,column=6).alignment = Alignment(horizontal='center')

                        #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=7).value = objetivo.signo
            ws.cell(row=cont,column=7).border = thin_border
            ws.cell(row=cont,column=7).alignment = Alignment(horizontal='center')

                        #agregamos la data, borde, alineacion
            if objetivo.operacion ==1:
                ws.cell(row=cont,column=8).value = 'Ponderado'
            elif objetivo.operacion ==2:
                ws.cell(row=cont,column=8).value = 'Promedio'         
            else:
                ws.cell(row=cont,column=8).value = 'Suma'
            ws.cell(row=cont,column=8).border = thin_border
            ws.cell(row=cont,column=8).alignment = Alignment(horizontal='center')

                        #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=9).value = objetivo.ponderacion
            ws.cell(row=cont,column=9).border = thin_border
            ws.cell(row=cont,column=9).alignment = Alignment(horizontal='center')

                        #agregamos la data, borde, alineacion
            if objetivo.periodos ==1:
                ws.cell(row=cont,column=10).value = 'Anual'
            elif objetivo.periodos ==2:
                ws.cell(row=cont,column=10).value = 'Semestral'
            elif objetivo.periodos ==3:
                ws.cell(row=cont,column=10).value = 'Quimestral'
            elif objetivo.periodos ==4:
                ws.cell(row=cont,column=10).value = 'Trimestral'
            else:
                ws.cell(row=cont,column=10).value = 'Mensual'
            ws.cell(row=cont,column=10).border = thin_border
            ws.cell(row=cont,column=10).alignment = Alignment(horizontal='center')

                        #agregamos la data, borde, alineacion
            if objetivo.meta ==1:
                ws.cell(row=cont,column=11).value = 'SI'
            else:
                ws.cell(row=cont,column=11).value = 'NO'
            ws.cell(row=cont,column=11).border = thin_border
            ws.cell(row=cont,column=11).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            if objetivo.observacion:
                ws.cell(row=cont,column=12).value = objetivo.observacion
            else:
                ws.cell(row=cont,column=12).value = '---'
            ws.cell(row=cont,column=12).border = thin_border
            ws.cell(row=cont,column=12).alignment = Alignment(horizontal='left')

            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="excel_objetivos.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 20.0
        ws.column_dimensions["D"].width = 40.0
        ws.column_dimensions["E"].width = 30.0
        ws.column_dimensions["F"].width = 25.0
        ws.column_dimensions["G"].width = 15.0
        ws.column_dimensions["H"].width = 25.0
        ws.column_dimensions["I"].width = 25.0
        ws.column_dimensions["J"].width = 15.0
        ws.column_dimensions["K"].width = 15.0
        ws.column_dimensions["L"].width = 25.0
        wb.save(response)
        return response


#GENERAR PDF
def PdfObjetivos(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        rows = Objetivos.objects.all().filter(estado=1)

        context = {
            'objetivo_list': rows,
        }

        pdf = render_to_pdf('reportes/pdf/pdfobjetivo.html', context)


            #FORZAR DOWNLOAD PDF
        if pdf:

            response = HttpResponse(pdf, content_type='application/pdf')

            filename = "Listado_objetivo_%s.pdf" %("_001")
            content = "inline; filename='%s'" %(filename)
            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content

            return response


def IngresarDepartamento(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            nombre = request.POST.get('nombre')
            if int(request.POST.get('departamentoPadre')) == 0:
                claveForanea_Departamento = None
            else:
                claveForanea_Departamento = Departamentos.objects.get(pk=request.POST.get('departamentoPadre'))
            if 'descripcion' in request.POST:
                descripcion = request.POST.get('descripcion')
            else:
                descripcion = None
            estado = request.POST.get('estado')

            departamentos = Departamentos.objects.filter(nombre=nombre).count()
            if departamentos == 0:
                with transaction.atomic():
                    departamento = Departamentos(
                        nombre = nombre,
                        descripcion = descripcion,
                        parent = claveForanea_Departamento
                    )
                    departamento.save()
                messages.add_message(request, messages.SUCCESS, 'Registro creado existosamente !!')
                return redirect('catalogos')
            else:
                messages.add_message(request, messages.WARNING, 'Departamento existente no puede ser creado !!')
                return redirect('catalogos')
        else:
            return redirect('logout')


def GridDepartamentosPadre(request):
    cursor = connection.cursor()
    query = " SELECT * FROM indicadores_departamentos WHERE indicadores_departamentos.parent_id IS NULL  "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def SubGridDepartamentosHijo(request):
    cursor = connection.cursor()
    query = " SELECT * FROM indicadores_departamentos WHERE indicadores_departamentos.parent_id =%s  "
    params = [request.GET['departamentoPadre_id']]
    cursor.execute(query, params)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def ConsultarDepartamento(request):
    departamento = Departamentos.objects.values().all().filter(pk=request.POST.get('departamento_id'))
    departamento_list = list(departamento)
    data = {
        'resultado':'ok_select',
        'departamento':departamento_list
    }

    return JsonResponse(data, safe=False)


def ActualizarDepartamento(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            departamentoId = request.POST.get('departamentoId')
            nombre = request.POST.get('nombre')
            if int(request.POST.get('departamentoPadre')) == 0:
                claveForanea_Objetivo = None
            else:
                claveForanea_Objetivo = Departamentos.objects.get(pk=request.POST.get('departamentoPadre'))
            if 'descripcion' in request.POST:
                descripcion = request.POST.get('descripcion')
            else:
                descripcion = None
            estado = request.POST.get('estado')

            with transaction.atomic():
                Departamentos.objects.filter(pk=departamentoId).update(
                    nombre = nombre,
                    parent = claveForanea_Objetivo,
                    descripcion = descripcion,
                    estado = estado,
                )
            messages.add_message(request, messages.SUCCESS, 'Departamento Actualizado.')
            return redirect('catalogos')


def ExcelDepartamentos(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        # consulta objetivos
        rows = Departamentos.objects.all().filter(estado=1)
        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Departamentos"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Listado de Departamentos'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Descripcion'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Dpto. Padre'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3 = ws['E3']
        e3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for departamento in rows:
            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=3).value = departamento.nombre
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=4).value = departamento.descripcion
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            # agregamos la data, borde, alineacion
            if departamento.parent:
                ws.cell(row=cont, column=5).value = departamento.parent.nombre
            else:
                ws.cell(row=cont, column=5).value = '-----'
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=5).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        # Establecemos el nombre del archivo
        nombre_archivo = "departamentos.xlsx"
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        # ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 50.0
        ws.column_dimensions["E"].width = 50.0
        wb.save(response)
        return response


def IngresarProducto(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            nombre = request.POST.get('nombre')
            if 'descripcion' in request.POST:
                descripcion = request.POST.get('descripcion')
            else:
                descripcion = None
            estado = request.POST.get('estado')

            productos = Productos.objects.filter(nombre=nombre).count()
            if productos == 0:
                with transaction.atomic():
                    producto = Productos(
                        nombre = nombre,
                        descripcion = descripcion,
                        estado = estado
                    )
                    producto.save()
                messages.add_message(request, messages.SUCCESS, 'Registro creado existosamente !!')
                return redirect('catalogos')
            else:
                messages.add_message(request, messages.WARNING, 'Producto/Servicio existente no puede ser creado !!')
                return redirect('catalogos')
        else:
            return redirect('logout')


def ConsultarProducto(request):
    producto = Productos.objects.values().all().filter(pk=request.POST.get('producto_id'))
    producto_list = list(producto)
    data = {
        'resultado':'ok_select',
        'producto':producto_list
    }

    return JsonResponse(data, safe=False)


def ActualizarProducto(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            productoId = request.POST.get('productoId')
            nombre = request.POST.get('nombre')
            if 'descripcion' in request.POST:
                descripcion = request.POST.get('descripcion')
            else:
                descripcion = None
            estado = request.POST.get('estado')

            with transaction.atomic():
                Productos.objects.filter(pk=productoId).update(
                    nombre = nombre,
                    descripcion = descripcion,
                    estado = estado,
                )
            messages.add_message(request, messages.SUCCESS, 'Producto/Servicio actualizado exitosamente.')
            return redirect('catalogos')


def ExcelProductos(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        # consulta objetivos
        rows = Productos.objects.all().filter(estado=1)
        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Productos o Servicios"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:D2')
        ws['B2'] = 'Listado de Productos/Servicios'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Descripcion'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for departamento in rows:
            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=3).value = departamento.nombre
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=4).value = departamento.descripcion
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        nombre_archivo = "productos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 50.0
        ws.column_dimensions["E"].width = 60.0
        wb.save(response)
        return response


def GridProductos(request):
    cursor = connection.cursor()
    query = " SELECT * FROM indicadores_productos "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def IngresarResponsable(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            nombre = request.POST.get('nombre')
            if 'correo' in request.POST:
                correo = request.POST.get('correo')
            else:
                correo = None
            if 'telefono' in request.POST:
                telefono = request.POST.get('telefono')
            else:
                telefono = None
            estado = request.POST.get('estado')

            responsables = Responsables.objects.filter(nombre=nombre).count()
            if responsables == 0:
                with transaction.atomic():
                    responsable = Responsables(
                        nombre = nombre,
                        correo = correo,
                        telefono = telefono,
                        estado = estado
                    )
                    responsable.save()
                messages.add_message(request, messages.SUCCESS, 'Registro creado existosamente !!')
                return redirect('catalogos')
            else:
                messages.add_message(request, messages.WARNING, 'Responsable existente no puede ser creado !!')
                return redirect('catalogos')
        else:
            return redirect('logout')


def GridResponsables(request):
    cursor = connection.cursor()
    query = " SELECT * FROM indicadores_responsables "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def ConsultarResponsable(request):
    responsable = Responsables.objects.values().all().filter(pk=request.POST.get('responsable_id'))
    responsable_list = list(responsable)
    data = {
        'resultado':'ok_select',
        'responsable':responsable_list
    }
    return JsonResponse(data, safe=False)


def ActualizarResponsable(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            responsableId = request.POST.get('responsableId')
            nombre = request.POST.get('nombre')
            if 'correo' in request.POST:
                correo = request.POST.get('correo')
            else:
                correo = None
            if 'telefono' in request.POST:
                telefono = request.POST.get('telefono')
            else:
                telefono = None
            estado = request.POST.get('estado')

            with transaction.atomic():
                Responsables.objects.filter(pk=responsableId).update(
                    nombre = nombre,
                    correo = correo,
                    telefono = telefono,
                    estado = estado
                )
            messages.add_message(request, messages.SUCCESS, 'Responsable actualizado exitosamente.')
            return redirect('catalogos')



def ExcelResponsables(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        rows = Responsables.objects.all().filter(estado=1)
        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Responsables"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Listado de Responsables'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Correo'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Telefono'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3 = ws['E3']
        e3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for responsable in rows:
            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=3).value = responsable.nombre
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=4).value = responsable.correo
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=5).value = responsable.telefono
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=5).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        nombre_archivo = "responsables.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 50.0
        ws.column_dimensions["E"].width = 60.0
        wb.save(response)
        return response


def IngresarPlan(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            nombre = request.POST.get('nombre')
            if 'descripcion' in request.POST:
                descripcion = request.POST.get('descripcion')
            else:
                descripcion = None
            estado = request.POST.get('estado')

            planes = Planes.objects.filter(nombre=nombre).count()
            if planes == 0:
                with transaction.atomic():
                    plan = Planes(
                        nombre = nombre,
                        descripcion = descripcion,
                        estado = estado
                    )
                    plan.save()
                messages.add_message(request, messages.SUCCESS, 'Registro creado existosamente !!')
                return redirect('catalogos')
            else:
                messages.add_message(request, messages.WARNING, 'Plan existente no puede ser creado !!')
                return redirect('catalogos')
        else:
            return redirect('logout')


def GridPlanes(request):
    cursor = connection.cursor()
    query = " SELECT * FROM indicadores_planes "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)



def ConsultarPlanDeAccion(request):
    plan = Planes.objects.values().all().filter(pk=request.POST.get('plan_id'))
    plan_list = list(plan)
    data = {
        'resultado':'ok_select',
        'plan':plan_list
    }
    return JsonResponse(data, safe=False)


def ActualizarPlan(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            planId = request.POST.get('planId')
            nombre = request.POST.get('nombre')
            if 'descripcion' in request.POST:
                descripcion = request.POST.get('descripcion')
            else:
                descripcion = None
            estado = request.POST.get('estado')

            with transaction.atomic():
                Planes.objects.filter(pk=planId).update(
                    nombre = nombre,
                    descripcion = descripcion,
                    estado = estado
                )
            messages.add_message(request, messages.SUCCESS, 'Plan actualizado exitosamente.')
            return redirect('catalogos')


def ExcelPlanesDeAccion(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        rows = Planes.objects.all().filter(estado=1)
        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Planes de Accion"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:D2')
        ws['B2'] = 'Planes de Accion'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Descripcion'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for plan in rows:
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=3).value = plan.nombre
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=4).value = plan.descripcion
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        nombre_archivo = "planesDeAccion.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.column_dimensions["C"].width = 60.0
        ws.column_dimensions["D"].width = 60.0
        wb.save(response)
        return response


def IngresarVariables(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            nombre = request.POST.get('nombre')
            operacion = request.POST.get('operacion')
            operacionEntrePeriodos = request.POST.get('operacionEntrePeriodos')
            estado = request.POST.get('estado')


            variables = Variables.objects.filter(nombre=nombre).count()
            if variables == 0:
                with transaction.atomic():
                    variable = Variables(
                        nombre = nombre,
                        operacion = operacion,
                        operacionEntrePeriodos = operacionEntrePeriodos,
                        estado = estado
                    )
                    variable.save()
                messages.add_message(request, messages.SUCCESS, 'Variable creada existosamente !!')
                return redirect('catalogos')
            else:
                messages.add_message(request, messages.WARNING, 'Variable existente no puede ser creado !!')
                return redirect('catalogos')
        else:
            return redirect('logout')


def GridVariables(request):
    cursor = connection.cursor()
    query = " SELECT * FROM indicadores_variables "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def ConsultarVariable(request):
    variable = Variables.objects.values().all().filter(pk=request.POST.get('variable_id'))
    variable_list = list(variable)
    data = {
        'resultado':'ok_select',
        'variable':variable_list
    }
    return JsonResponse(data, safe=False)


def ActualizarVariable(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            variableId = request.POST.get('variableId')
            nombre = request.POST.get('nombre')
            operacion = request.POST.get('operacion')
            operacionEntrePeriodos = request.POST.get('operacionEntrePeriodos')
            estado = request.POST.get('estado')

            with transaction.atomic():
                Variables.objects.filter(pk=variableId).update(
                    nombre = nombre,
                    operacion = operacion,
                    operacionEntrePeriodos = operacionEntrePeriodos,
                    estado = estado
                )
            messages.add_message(request, messages.SUCCESS, 'Variable actualizada exitosamente.')
            return redirect('catalogos')



def ExcelVariables(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        rows = Variables.objects.all().filter(estado=1)
        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Variables"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Variables'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Operacion'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Operacion entre Periodos'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3 = ws['E3']
        e3.font = Font(bold=True, color=colors.DARKBLUE, size=12)


        cont = 4
        indice = 1
        for variable in rows:
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=3).value = variable.nombre
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            if variable.operacion == 1:
                operacion = 'Ponderado'
            elif variable.operacion == 2:
                operacion = 'Promedio'
            elif variable.operacion == 3:
                operacion = 'Suma'

            ws.cell(row=cont, column=4).value = operacion
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            if variable.operacionEntrePeriodos == 1:
                operacionEntrePeriodos = 'Ponderado'
            elif variable.operacionEntrePeriodos == 2:
                operacionEntrePeriodos = 'Promedio'
            elif variable.operacionEntrePeriodos == 3:
                operacionEntrePeriodos = 'Suma'

            ws.cell(row=cont, column=5).value = operacionEntrePeriodos
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=5).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        nombre_archivo = "variables.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.column_dimensions["C"].width = 60.0
        ws.column_dimensions["D"].width = 30.0
        ws.column_dimensions["E"].width = 40.0
        wb.save(response)
        return response


def IngresarFormula(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            estado = request.POST.get('estado')

            if 'operador1' in request.POST:
                operador_1 = request.POST.get('operador1')
            else:
                operador_1 = None
            if int(request.POST.get('variable1')) > 0:
                claveForanea_variable1 = Variables.objects.get(pk=request.POST.get('variable1'))
            else:
                claveForanea_variable1 = None

            if 'operador2' in request.POST:
                operador_2 = request.POST.get('operador2')
            else:
                operador_2 = None
            if 'variable2' in request.POST:
                if int(request.POST.get('variable2')) > 0:
                    claveForanea_variable2 = Variables.objects.get(pk=request.POST.get('variable2'))
            else:
                claveForanea_variable2 = None

            if 'operador3' in request.POST:
                operador_3 = request.POST.get('operador3')
            else:
                operador_3 = None
            if 'variable3' in request.POST:
                if int(request.POST.get('variable3')) > 0:
                    claveForanea_variable3 = Variables.objects.get(pk=request.POST.get('variable3'))
            else:
                claveForanea_variable3 = None

            if 'operador4' in request.POST:
                operador_4 = request.POST.get('operador4')
            else:
                operador_4 = None
            if 'variable4' in request.POST:
                if int(request.POST.get('variable4')) > 0:
                    claveForanea_variable4 = Variables.objects.get(pk=request.POST.get('variable4'))
            else:
                claveForanea_variable4 = None

            if 'operador5' in request.POST:
                operador_5 = request.POST.get('operador5')
            else:
                operador_5 = None
            if 'variable5' in request.POST:
                if int(request.POST.get('variable5')) > 0:
                    claveForanea_variable5 = Variables.objects.get(pk=request.POST.get('variable5'))
            else:
                claveForanea_variable5 = None

            if 'operador6' in request.POST:
                operador_6 = request.POST.get('operador6')
            else:
                operador_6 = None
            if 'variable6' in request.POST:
                if int(request.POST.get('variable6')) > 0:
                    claveForanea_variable6 = Variables.objects.get(pk=request.POST.get('variable6'))
            else:
                claveForanea_variable6 = None

            estado = request.POST.get('estado')
            formulas = Formulas.objects.filter(nombre=nombre).count()
            if formulas == 0:
                with transaction.atomic():
                    formula = Formulas(
                        nombre = nombre,
                        descripcion = descripcion,
                        operador1 = operador_1,
                        variable1 = claveForanea_variable1,
                        operador2 = operador_2,
                        variable2 = claveForanea_variable2,
                        operador3 = operador_3,
                        variable3 = claveForanea_variable3,
                        operador4 = operador_4,
                        variable4 = claveForanea_variable4,
                        operador5 = operador_5,
                        variable5 = claveForanea_variable5,
                        operador6 = operador_6,
                        variable6 = claveForanea_variable6,
                        estado = estado
                    )
                    formula.save()
                messages.add_message(request, messages.SUCCESS, 'Fórmula creada existosamente !!')
                return redirect('catalogos')
            else:
                messages.add_message(request, messages.WARNING, 'Fórmula existente no puede ser creado !!')
                return redirect('catalogos')
        else:
            return redirect('logout')


def GridFormulas(request):
    cursor = connection.cursor()
    query = " SELECT indicadores_formulas.id, indicadores_formulas.nombre, indicadores_formulas.descripcion, "
    query += " indicadores_formulas.operador1, indicadores_variables.nombre AS variable_1, indicadores_formulas.operador2, "
    query += " variable_2.nombre AS variable_2, indicadores_formulas.operador3, variable_3.nombre AS variable_3, "
    query += " indicadores_formulas.operador4, variable_4.nombre AS variable_4, indicadores_formulas.operador5, "
    query += " variable_5.nombre AS variable_5, indicadores_formulas.operador6, variable_6.nombre AS variable_6 "
    query += " FROM indicadores_formulas "
    query += " LEFT JOIN indicadores_variables ON indicadores_formulas.variable1_id = indicadores_variables.id "
    query += " LEFT JOIN indicadores_variables variable_2 ON indicadores_formulas.variable2_id = variable_2.id "
    query += " LEFT JOIN indicadores_variables variable_3 ON indicadores_formulas.variable3_id = variable_3.id "
    query += " LEFT JOIN indicadores_variables variable_4 ON indicadores_formulas.variable4_id = variable_4.id "
    query += " LEFT JOIN indicadores_variables variable_5 ON indicadores_formulas.variable5_id = variable_5.id "
    query += " LEFT JOIN indicadores_variables variable_6 ON indicadores_formulas.variable6_id = variable_6.id "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def ConsultarFormula(request):
    formula = Formulas.objects.values().all().filter(pk=request.POST.get('formula_id'))
    formula_list = list(formula)
    data = {
        'resultado':'ok_select',
        'formula':formula_list
    }
    return JsonResponse(data, safe=False)


def ActualizarFormula(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            formulaId = request.POST.get('formulaId')
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            estado = request.POST.get('estado')

            if 'operador1' in request.POST:
                operador_1 = request.POST.get('operador1')
            else:
                operador_1 = None
            if int(request.POST.get('variable1')) > 0:
                claveForanea_variable1 = Variables.objects.get(pk=request.POST.get('variable1'))
            else:
                claveForanea_variable1 = None

            if 'operador2' in request.POST:
                operador_2 = request.POST.get('operador2')
            else:
                operador_2 = None
            if 'variable2' in request.POST:
                if int(request.POST.get('variable2')) > 0:
                    claveForanea_variable2 = Variables.objects.get(pk=request.POST.get('variable2'))
            else:
                claveForanea_variable2 = None

            if 'operador3' in request.POST:
                operador_3 = request.POST.get('operador3')
            else:
                operador_3 = None
            if 'variable3' in request.POST:
                if int(request.POST.get('variable3')) > 0:
                    claveForanea_variable3 = Variables.objects.get(pk=request.POST.get('variable3'))
            else:
                claveForanea_variable3 = None

            if 'operador4' in request.POST:
                operador_4 = request.POST.get('operador4')
            else:
                operador_4 = None
            if 'variable4' in request.POST:
                if int(request.POST.get('variable4')) > 0:
                    claveForanea_variable4 = Variables.objects.get(pk=request.POST.get('variable4'))
            else:
                claveForanea_variable4 = None

            if 'operador5' in request.POST:
                operador_5 = request.POST.get('operador5')
            else:
                operador_5 = None
            if 'variable5' in request.POST:
                if int(request.POST.get('variable5')) > 0:
                    claveForanea_variable5 = Variables.objects.get(pk=request.POST.get('variable5'))
            else:
                claveForanea_variable5 = None

            if 'operador6' in request.POST:
                operador_6 = request.POST.get('operador6')
            else:
                operador_6 = None
            if 'variable6' in request.POST:
                if int(request.POST.get('variable6')) > 0:
                    claveForanea_variable6 = Variables.objects.get(pk=request.POST.get('variable6'))
            else:
                claveForanea_variable6 = None


            with transaction.atomic():
                Formulas.objects.filter(pk=formulaId).update(
                    nombre = nombre,
                    descripcion = descripcion,
                    operador1 = operador_1,
                    variable1 = claveForanea_variable1,
                    operador2 = operador_2,
                    variable2 = claveForanea_variable2,
                    operador3 = operador_3,
                    variable3 = claveForanea_variable3,
                    operador4 = operador_4,
                    variable4 = claveForanea_variable4,
                    operador5 = operador_5,
                    variable5 = claveForanea_variable5,
                    operador6 = operador_6,
                    variable6 = claveForanea_variable6,
                    estado = estado
                )
                messages.add_message(request, messages.SUCCESS, 'Fórmula actualizada existosamente !!')
                return redirect('catalogos')


def ExcelFormulas(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        rows = Formulas.objects.all().filter(estado=1)
        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Formulas"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Formulas'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Descripción'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Fórmula Resultante'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3 = ws['E3']
        e3.font = Font(bold=True, color=colors.DARKBLUE, size=12)


        cont = 4
        indice = 1
        for formula in rows:
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=3).value = formula.nombre
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=4).value = formula.descripcion
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            formulaResultante = formula.operador1 + str(formula.variable1.nombre)

            formulaResultante += formula.operador2
            if formula.variable2:
                formulaResultante += str(formula.variable2.nombre)

            formulaResultante += formula.operador3
            if formula.variable3:
                formulaResultante += str(formula.variable3.nombre)

            formulaResultante += formula.operador4
            if formula.variable4:
                formulaResultante += str(formula.variable4.nombre)

            formulaResultante += formula.operador5
            if formula.variable5:
                formulaResultante += str(formula.variable5.nombre)

            formulaResultante += formula.operador6
            if formula.variable6:
                formulaResultante += str(formula.variable6.nombre)


            ws.cell(row=cont, column=5).value = formulaResultante
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=5).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        nombre_archivo = "formulas.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.column_dimensions["C"].width = 50.0
        ws.column_dimensions["D"].width = 40.0
        ws.column_dimensions["E"].width = 50.0
        wb.save(response)
        return response


def IngresarEjeEstrategico(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            estado = request.POST.get('estado')

            ejes = EjesEstrategicos.objects.filter(nombre=nombre).count()
            if ejes == 0:
                with transaction.atomic():
                    ejeEstrategico = EjesEstrategicos(
                        nombre = nombre,
                        descripcion = descripcion,
                        estado = estado
                    )
                    ejeEstrategico.save()
                messages.add_message(request, messages.SUCCESS, 'Eje creado existosamente !!')
                return redirect('catalogos')
            else:
                messages.add_message(request, messages.WARNING, 'Eje existente no puede ser creado !!')
                return redirect('catalogos')
        else:
            return redirect('logout')


def GridEjesEstrategicos(request):
    cursor = connection.cursor()
    query = " SELECT * FROM indicadores_ejesestrategicos "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def ConsultarEjesEstrategicos(request):
    eje = EjesEstrategicos.objects.values().all().filter(pk=request.POST.get('ejeestrategico_id'))
    eje_list = list(eje)
    data = {
        'resultado':'ok_select',
        'ejesestrategico':eje_list
    }
    return JsonResponse(data, safe=False)


def ActualizarEjeEstrategico(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            ejeId = request.POST.get('ejeId')
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            estado = request.POST.get('estado')

            with transaction.atomic():
                EjesEstrategicos.objects.filter(pk=ejeId).update(
                    nombre = nombre,
                    descripcion = descripcion,
                    estado = estado
                )
            messages.add_message(request, messages.SUCCESS, 'Eje actualizado exitosamente.')
            return redirect('catalogos')


def ExcelEjesEstrategicos(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        rows = EjesEstrategicos.objects.all().filter(estado=1)
        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Ejes Estrategicos"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:D2')
        ws['B2'] = 'Ejes Estratégicos'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Descripción'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for eje in rows:
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=3).value = eje.nombre
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=4).value = eje.descripcion
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        nombre_archivo = "eje.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.column_dimensions["C"].width = 40.0
        ws.column_dimensions["D"].width = 40.0
        wb.save(response)
        return response


def IngresarPesoResponsable(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST':
            responsable = request.POST.get('responsable')
            claveForanea_Responsable= Responsables.objects.get(pk=request.POST.get('responsable'))
            departamento = request.POST.get('departamento')
            claveForanea_Departamento = Departamentos.objects.get(pk=request.POST.get('departamento'))
            pesoEmpresa = request.POST.get('pesoEmpresa')
            pesoIndividual = request.POST.get('pesoIndividual')
            pesoTrabajoEquipo = request.POST.get('pesoTrabajoEquipo')
            pesoOtros = request.POST.get('pesoOtros')
            estado = request.POST.get('estado')

            with transaction.atomic():
                pesoResponsable = PesosResponsables(
                    responsable = claveForanea_Responsable,
                    departamento = claveForanea_Departamento,
                    pesoEmpresa = pesoEmpresa,
                    pesoIndividual = pesoIndividual,
                    pesoTrabajoEquipo = pesoTrabajoEquipo,
                    pesoOtros = pesoOtros,
                    estado = estado
                )
                pesoResponsable.save()
                messages.add_message(request, messages.SUCCESS, 'Peso Creado Existosamente !!')
                return redirect('catalogos')
        else:
            return redirect('logout')


def GridPesosResponsables(request):
    cursor = connection.cursor()
    query = " SELECT indicadores_responsables.nombre AS responsable, indicadores_departamentos.nombre AS departamento, indicadores_pesosresponsables.pesoEmpresa, "
    query += " indicadores_pesosresponsables.pesoIndividual, indicadores_pesosresponsables.pesoTrabajoEquipo, indicadores_pesosresponsables.pesoOtros, "
    query += " indicadores_pesosresponsables.estado, indicadores_pesosresponsables.id "
    query += " FROM indicadores_pesosresponsables "
    query += " LEFT JOIN indicadores_responsables ON indicadores_pesosresponsables.responsable_id = indicadores_responsables.id "
    query += " LEFT JOIN indicadores_departamentos ON indicadores_pesosresponsables.departamento_id = indicadores_departamentos.id "
    cursor.execute(query)
    rows = dictfetchall(cursor)

    return JsonResponse(rows, safe=False)


def ConsultarPesosResponsables(request):
    pesoresponsable = PesosResponsables.objects.values().all().filter(pk=request.POST.get('pesoresponsable_id'))
    pesoresponsable_list = list(pesoresponsable)
    data = {
        'resultado':'ok_select',
        'pesoresponsable':pesoresponsable_list
    }
    return JsonResponse(data, safe=False)


def ActualizarPesoResponsable(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method=='POST':
            pesoResponsableId = request.POST.get('pesoResponsableId')
            responsable = request.POST.get('responsable')
            claveForanea_Responsable= Responsables.objects.get(pk=request.POST.get('responsable'))
            departamento = request.POST.get('departamento')
            claveForanea_Departamento = Departamentos.objects.get(pk=request.POST.get('departamento'))
            pesoEmpresa = request.POST.get('pesoEmpresa')
            pesoIndividual = request.POST.get('pesoIndividual')
            pesoTrabajoEquipo = request.POST.get('pesoTrabajoEquipo')
            pesoOtros = request.POST.get('pesoOtros')
            estado = request.POST.get('estado')

            with transaction.atomic():
                PesosResponsables.objects.filter(pk=pesoResponsableId).update(
                    responsable = claveForanea_Responsable,
                    departamento = claveForanea_Departamento,
                    pesoEmpresa = pesoEmpresa,
                    pesoIndividual = pesoIndividual,
                    pesoTrabajoEquipo = pesoTrabajoEquipo,
                    pesoOtros = pesoOtros,
                    estado = estado
                )
            messages.add_message(request, messages.SUCCESS, 'Peso actualizado exitosamente.')
            return redirect('catalogos')


def ExcelPesosResponsables(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        rows = PesosResponsables.objects.all().filter(estado=1)
        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        # Creamos el libro de trabajo
        wb = Workbook()
        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Pesos Responsables"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:H2')
        ws['B2'] = 'Pesos Responsables'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Responsable'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Departamento'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Peso Empresa'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3 = ws['E3']
        e3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['F3'] = 'Peso Individual'
        ws.cell(row=3, column=6).border = thin_border
        ws.cell(row=3, column=6).alignment = Alignment(horizontal='center')
        f3 = ws['F3']
        f3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['G3'] = 'Peso Equipo Trabajo'
        ws.cell(row=3, column=7).border = thin_border
        ws.cell(row=3, column=7).alignment = Alignment(horizontal='center')
        g3 = ws['G3']
        g3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['H3'] = 'Otros Pesos'
        ws.cell(row=3, column=8).border = thin_border
        ws.cell(row=3, column=8).alignment = Alignment(horizontal='center')
        h3 = ws['H3']
        h3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for peso in rows:
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=3).value = peso.responsable.nombre
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=4).value = peso.departamento.nombre
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=5).value = peso.pesoEmpresa
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=5).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=6).value = peso.pesoIndividual
            ws.cell(row=cont, column=6).border = thin_border
            ws.cell(row=cont, column=6).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=7).value = peso.pesoTrabajoEquipo
            ws.cell(row=cont, column=7).border = thin_border
            ws.cell(row=cont, column=7).alignment = Alignment(horizontal='center')

            ws.cell(row=cont, column=8).value = peso.pesoOtros
            ws.cell(row=cont, column=8).border = thin_border
            ws.cell(row=cont, column=8).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        nombre_archivo = "eje.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 30.0
        ws.column_dimensions["E"].width = 20.0
        ws.column_dimensions["F"].width = 20.0
        ws.column_dimensions["G"].width = 25.0
        ws.column_dimensions["H"].width = 20.0
        ws.column_dimensions["I"].width = 20.0
        wb.save(response)
        return response




def Planificacion(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        context={}
        return render(request, 'indicadores/planificacion.html', context)




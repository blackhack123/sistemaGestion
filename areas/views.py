# -*- coding: utf-8 -*-
from django.shortcuts import render, redirect
from areas.models import Areas, Procesos
from personal.models import Personal, Cargo, AreaPersonal
from django.http import HttpResponse, JsonResponse
import json
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#IMPORTAMOS BORDES 
from openpyxl.styles import Border, Side, Alignment, Color, Fill, colors, Font
from openpyxl.cell import Cell
#importamos render_to_pdf de vista principal
from principal.views import render_to_pdf
#IMPORTAMOS LA CONEXION PARA QUERY PERSONALIZADO
from django.db import connection
from principal.views import dictfetchall

#impor andres
from personal.forms import AreaForm
from django.contrib import messages


# Create your views here.

def areas(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        context={
            'cargos':Cargo.objects.values('id', 'nombre').all().filter(estado=1),
            #'personal':Personal.objects.values('id', 'nombre').all().filter(estado=1),
        }

        return render(request, 'admin/areas.html', context)


def selectJefe(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_cargo=request.POST.get('id_cargo')
            jefes=Personal.objects.values('id', 'nombre').all().filter(estado=1, id_cargo_id=id_cargo)
            jefes_list = list(jefes)
            
            data = {
            'resultado':'ok_select',
            'jefes_list': jefes_list,

            }
            return HttpResponse(json.dumps(data), content_type="application/json")




def gridAreas(request):

        
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        cursor = connection.cursor()
        query = " SELECT areas_areas.id, areas_areas.area,  areas_areas.descripcion, areas_areas.estado, "
        query = query + " personal_personal.nombre AS encargado, areas_areas.tipo_proceso"
        query = query + " FROM areas_areas "
        query = query + " LEFT JOIN personal_personal ON areas_areas.id_personal_id = personal_personal.id ORDER BY area "
        cursor.execute(query)
        rows = dictfetchall(cursor)

        if rows:
            #RETORNA  
            return JsonResponse(rows, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Areas',
            }
            return JsonResponse(json, safe=False)
    

def selectArea(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_area = request.POST.get('id_area')

            area = Areas.objects.get(pk=id_area)

            area_list ={
                'id':area.id,
                'nombre':area.area ,
                'descripcion':area.descripcion,
                'id_personal':area.id_personal_id,
                'estado':area.estado,
                'tipo_proceso':area.tipo_proceso
            }
            
            personal=Personal.objects.values('id', 'nombre','id_cargo_id').all().filter(pk=area.id_personal_id)
            personal_list=list(personal)
            data = {
            'resultado':'ok_select',
            'area_list':area_list,
            'personal':personal_list,
            }
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})

def insertArea(request):


    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            id_personal = request.POST.get('id_personal')
            estado = request.POST.get('estado')
            tipoProceso=request.POST.get('tipoProceso')
            fec_ingresa = request.POST.get('datetime')
            usuario_ing_id=request.session['idUsuario']
            try:
                #consultar  nombre
                area = Areas.objects.get(area=nombre) 
                area_list ={
                    'idUsuario':area.id,
                    'nombre':area.area ,
                    'descripcion':area.descripcion,
                    'id_personal':area.id_personal_id,
                    'estado':area.estado,
                    'tipo_proceso':area.tipo_proceso
                }
                data = {
                    'resultado':'no_ok',
                    'mensaje': 'Area existente desea modificarlo ?',
                    'area_list':area_list,
                }
                return HttpResponse(json.dumps(data), content_type="application/json")

            except Areas.DoesNotExist :

                #ASIGNAMOS DATOS AL OBJETO
                area=Areas(
                    area= nombre, 
                    descripcion= descripcion, 
                    id_personal_id= id_personal,
                    estado=estado,
                    fec_ingresa=fec_ingresa,
                    usuario_ing_id=usuario_ing_id,
                    tipo_proceso = tipoProceso
                )
                #guardar datos
                area.save()

                #creamos el directorio para el proceso
                #importamos la libreria OS
                import os
                nombre_absoluto= "media/gestionDocumental/"+str(tipoProceso.encode('utf-8'))+"/"+str(nombre.encode('utf-8'))

                if not os.path.exists(nombre_absoluto):
		
                    os.makedirs(nombre_absoluto, mode=0777)

                #retornamos la respuesta
                data = {
                    'resultado':'ok_insert',
                    'mensaje': 'Datos Guardados Correctamente !!',
                }
                return HttpResponse(json.dumps(data), content_type="application/json")



def updateArea(request):

    if 'nombreUsuario' not in request.session:

            return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():


            id_area = request.POST.get('id_area')
            carpeta_anterior = request.POST.get('carpeta_anterior')
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            id_personal = request.POST.get('id_personal')
            estado = request.POST.get('estado')
            #fec_modifica=request.POST.get('datetime')
            usuario_mod_id=request.session['idUsuario']
            tipoProceso = request.POST.get('tipoProceso')

            #update
            Areas.objects.filter(pk=id_area).update(
                area= nombre, 
                descripcion= descripcion, 
                id_personal_id= id_personal,
                estado=estado,
                fec_modifica= None,
                usuario_mod_id=usuario_mod_id,
                tipo_proceso = tipoProceso
            )

            tipo_anterior = request.POST.get('tipo_anterior')
            path = 'media/gestionDocumental/'
            import os

            #valida si existe un cambio en el nombre de l procesos
            #renombra la carpeta
            if (carpeta_anterior != nombre) and (tipo_anterior == tipoProceso):
                try:
                    for root, dirs, files in os.walk(path):
                     
                        for directory in dirs:
                            if directory == carpeta_anterior:
                                os.rename(str(root+'/'+carpeta_anterior.encode('utf-8')), str(root+'/'+nombre.encode('utf-8')))
                except OSError:
                    pass

            #valida si existe un cambio en el tipo de proceso
            #muevo la carpeta al nuevo destino
            if (carpeta_anterior == nombre) and (tipo_anterior != tipoProceso):
                try:
                    for root, dirs, files in os.walk(path):
                        for directory in dirs:
                            if directory == nombre:
                                import shutil
                                from shutil import Error
                                shutil.move(str(root+'/'+nombre.encode('utf-8')), str('media/gestionDocumental'+'/'+tipoProceso.encode('utf-8')+'/'))
                except Error as error:
                    pass

            if (carpeta_anterior != nombre) and (tipo_anterior != tipoProceso):
                try:
                    for root, dirs, files in os.walk(path):
                     
                        for directory in dirs:
                            if directory == carpeta_anterior:
                                os.rename(str(root+'/'+carpeta_anterior.encode('utf-8')), str(root+'/'+nombre.encode('utf-8')))
                                import shutil
                                shutil.move(str(root+'/'+nombre.encode('utf-8')), str('media/gestionDocumental'+'/'+tipoProceso.encode('utf-8')+'/'))
                except OSError:
                    pass

            data = {
                'resultado':'ok_update',
                'mensaje': 'Datos Actualizados Correctamente !!',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")


def excelAreas(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        #Consultar las Bodegas
        cursor = connection.cursor()
        query = " SELECT areas_areas.id, areas_areas.area,  areas_areas.descripcion, areas_areas.estado, "
        query = query + " personal_personal.nombre AS encargado"
        query = query + " FROM areas_areas "
        query = query + " LEFT JOIN personal_personal ON areas_areas.id_personal_id = personal_personal.id "

        cursor.execute(query)
        rows = dictfetchall(cursor)

        areas_list = list(rows)

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Procesos"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Reporte General de Procesos'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Proceso'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Descripción'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Encargado de Proceso'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        cont = 4
        indice = 1
        for area in areas_list:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = area['area']
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=4).value = area['descripcion']
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=5).value = area['encargado']
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')

            if area['estado'] == "0":

                #obtengo celdas de estado inactivo
                areaCell=ws.cell(row=cont,column=3)
                descripcionCell=ws.cell(row=cont,column=4)
                encargadoCell=ws.cell(row=cont,column=5)

                #color rojo a inactivos
                areaCell.font = Font(color=colors.RED)
                descripcionCell.font = Font(color=colors.RED)
                encargadoCell.font = Font(color=colors.RED)

            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Reporte_General_proceso.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 40.0
        ws.column_dimensions["E"].width = 40.0
        wb.save(response)
        #retorna el archivo excel
        return response



def pdfAreas(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        #Consultar las Bodegas
        cursor = connection.cursor()
        query = " SELECT areas_areas.id, areas_areas.area,  areas_areas.descripcion, areas_areas.estado, "
        query = query + " personal_personal.nombre AS encargado"
        query = query + " FROM areas_areas "
        query = query + " LEFT JOIN personal_personal ON areas_areas.id_personal_id = personal_personal.id "

        cursor.execute(query)
        rows = dictfetchall(cursor)

        areas_list = list(rows)

        context = {
            'areas': areas_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfAreas.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Reporte_Areas_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response
	  

def selecAreapersonal(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        if request.method=='POST':
            idareapersonal = request.POST.get('idareapersonal')
        
            # consulta un areapersonal por su id
            areapersonal = AreaPersonal.objects.values('id','nombre','descripcion','estado').all().filter(pk=idareapersonal)
            areapersonal_lista = list(areapersonal)

            data={
                'resultado':'ok_select',
                'areapersonal':areapersonal_lista
            }
            return HttpResponse(json.dumps(data), content_type='application/json')


def ActualizarAreapersonal(request):
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        if request.method=='POST':
        
            idareapersonal = request.POST.get('id')
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            estado = request.POST.get('estado')

            AreaPersonal.objects.filter(pk=idareapersonal).update(
                nombre= nombre,
                descripcion=descripcion,
                estado=estado
            )

            messages.add_message(request, messages.SUCCESS, 'REGISTRO ACTUALIZADO')
            request.session['update_area'] = 'ok_update'
            return redirect('cargo')


def EliminarAreapersonal(request):
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        idareapersonal = request.POST.get('idareapersonal')
        #eliminar vehiculo de embarque
        AreaPersonal.objects.filter(pk=idareapersonal).delete()
        
        data={
            'resultado':'ok_delete',
        }
        return HttpResponse(json.dumps(data), content_type='application/json')


def CrearAreapersonal(request):
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        if request.method=='POST':

            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')             
            estado = request.POST.get('estado') 
            
            
            areapersonal = AreaPersonal.objects.filter(nombre=nombre).count()
            if areapersonal == 0:
                areapersonal=AreaPersonal(
                    nombre = nombre,
                    descripcion=descripcion,                  
                    estado=estado,                   
                )
                areapersonal.save()
                #mensaje exitoso
                messages.add_message(request, messages.SUCCESS, 'Registro Creado Existosamente !!')
                request.session['insert_area']='ok_insert'
                return redirect('cargo')
            else:
                request.session['insert_area'] = 'ok_insert'
                messages.add_message(request, messages.WARNING, 'El registro ya existe.. no puede ser creado !!')
                return redirect('cargo')
        else:
            return redirect('logout')



def ExcelAreapersonal(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        areasPersonal_lista = AreaPersonal.objects.all()

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Areas"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Reporte General de Areas'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Descripción'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Estado'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for area in areasPersonal_lista:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = area.nombre
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=4).value = area.descripcion
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')
          
            if area.estado == '1':
                ws.cell(row=cont,column=5).value = 'activo'
            else:
                ws.cell(row=cont,column=5).value = 'inactivo'
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')


            if area.estado == '0':

                #obtengo celdas de estado inactivo
                nombreCell=ws.cell(row=cont,column=3)
                descripcionCell=ws.cell(row=cont,column=4)
                estadoCell=ws.cell(row=cont,column=5)               

                #color rojo a inactivos
                nombreCell.font = Font(color=colors.RED)
                descripcionCell.font = Font(color=colors.RED)
                estadoCell.font = Font(color=colors.RED)
                
            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Reporte_General_areas.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 40.0
        ws.column_dimensions["E"].width = 40.0
        wb.save(response)
        #retorna el archivo excel
        return response


def PdfAreapersonal(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        #Consultar las Bodegas
        areasPersonal_list = AreaPersonal.objects.all()
        context = {
            'areasdos': areasPersonal_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfAreaspersonal.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Reporte_Areas_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response



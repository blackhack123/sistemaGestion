# -*- coding: utf-8 -*-
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
import os
import json, md5
#import datetime
from django.core import serializers
from datetime import datetime  
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#IMPORTAMOS BORDES 
from openpyxl.styles import Border, Side, Alignment, Color, Fill, colors, Font
from openpyxl.cell import Cell
#librerias para pdf
from io import BytesIO
from django.template import Context
from django.template.loader import get_template
from xhtml2pdf import pisa 
#IMPORTAMOS LA CONEXION PARA QUERY PERSONALIZADO
from django.db import connection
from django.db.models import Q

#importamos Usuarios
from seguridad.models import Usuario, TipoUsuario
from areas.models import Procesos
from personal.models import Personal
from sac.models import Sac, PlanAccion

#vistas
def home(request):
    
    if 'nombreUsuario' not in request.session:
        return redirect('logout')
        #return render(request, 'seguridad/login.html', {})

    else:
        cursor = connection.cursor()
        query = " SELECT personal_personal.id, personal_personal.nombre, personal_cargo.nombre AS cargo "
        query = query + " FROM personal_personal LEFT JOIN personal_cargo ON personal_personal.id_cargo_id = personal_cargo.id "
        query = query + " WHERE personal_personal.estado = 1 "
        cursor.execute(query)
        rows_personal = dictfetchall(cursor)
        context={
            #'sacAuditor':Sac.objects.values('estado_encabezado').all().filter(estado_encabezado=0),
            'niveles':TipoUsuario.objects.all(),
            'procesos':Procesos.objects.values('id', 'proceso').all().filter(estado=1),
            'personal':list(rows_personal),
            'usuarios':Usuario.objects.values('id', 'usuario').all().filter(estado=1),
        }
        return render(request, 'admin/index.html',context)


def render_to_pdf(template_src, context_dict={}):
    
     template = get_template(template_src)
     html  = template.render(context_dict)
     result = BytesIO()
     pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
     if not pdf.err:
         return HttpResponse(result.getvalue(), content_type='application/pdf')
     return None


def gridUsuarios(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:


        cursor = connection.cursor()
        query = " SELECT seguridad_usuario.usuario,  seguridad_usuario.estado, seguridad_usuario.id_tipo_usuario_id, seguridad_usuario.id, "
        query = query + " seguridad_usuario.id_personal_id, seguridad_usuario.director, seguridad_usuario.colaborador, seguridad_usuario.auditor, "
        query = query + " seguridad_usuario.auditor_lider, seguridad_usuario.lider_norma, personal_personal.nombre AS nombre_personal, personal_personal.telefono, "
        query = query + " personal_personal.correo, personal_personal.id_cargo_id, areas_areas.area AS proceso, seguridad_usuario.permiso_lectura, seguridad_usuario.permiso_escritura "
        query = query + " FROM seguridad_usuario "
        query = query + " LEFT JOIN personal_personal ON seguridad_usuario.id_personal_id = personal_personal.id "
        query = query + " LEFT JOIN areas_area_proceso ON personal_personal.id = areas_area_proceso.personal_id  "
        query = query + " LEFT JOIN areas_areas ON areas_area_proceso.area_id = areas_areas.id "
        query = query + " GROUP BY usuario "
        query = query + " ORDER BY nombre_personal "


        cursor.execute(query)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        usuarios_list = list(rows)

        if usuarios_list:
            #RETORNA  
            return JsonResponse(usuarios_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Usuarios',
            }
            return JsonResponse(json, safe=False)


def selectUsuario(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        # validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            idUsuario = request.POST.get('id_usuario')
            nombre = request.POST.get('nombre')

            # consulta ususario id y nombre de personal

            usuario = Usuario.objects.get(pk=idUsuario)
            personal = Personal.objects.get(nombre=nombre)

            usuario_list = {
                'idUsuario': usuario.id,
                'id_personal': personal.nombre,
                'usuario': usuario.usuario,
                'estado': usuario.estado,
                'nivel': usuario.id_tipo_usuario_id,
                'director':usuario.director,
                'colaborador':usuario.colaborador,
                'auditor': usuario.auditor,
                'auditor_lider': usuario.auditor_lider,
                'lider_norma': usuario.lider_norma,
                'permiso_lectura': usuario.permiso_lectura,
                'permiso_escritura': usuario.permiso_escritura
            }

            data = {
                'resultado': 'ok_select',
                'usuario_list': usuario_list,
            }
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})


def insertUsuarios(request):

    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        # validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():
            nombre = request.POST.get('id_personal')
            usuario = request.POST.get('usuario')
            clave = request.POST.get('clave')
            claveNueva = md5.new(clave).hexdigest()
            id_tipo_usuario = request.POST.get('id_tipo_usuario_id')
            estado = request.POST.get('estado')
            fec_ingresa = request.POST.get('datetime')
            usuario_ing_id = request.session['idUsuario']
            try:
                personal = Personal.objects.get(nombre=nombre)
            except Personal.DoesNotExist:
                data = {
                    'resultado': 'no_ok_personal',
                    'mensaje': 'No existe el Colaborador !!',
                }
                return HttpResponse(json.dumps(data), content_type="application/json")


            cursor = connection.cursor()
            query = " SELECT * "
            query = query + " FROM seguridad_usuario "
            query = query + " WHERE seguridad_usuario.usuario= %(usuario)s "
            query = query + " AND seguridad_usuario.id_personal_id= %(id_personal_id)s "
            query = query + " OR seguridad_usuario.usuario= %(usuario)s "

            params = {'usuario': usuario, 'id_personal_id': personal.id}

            cursor.execute(query, params)
            rows = dictfetchall(cursor)
            usuario_list = list(rows)
            total_elementos = len(usuario_list)

            if total_elementos >= 1:
                # consultar bodega por nombre
                usuario_list = {
                    'idUsuario': usuario_list[0]['id'],
                    'id_personal': personal.nombre,
                    # 'id_personal':usuario.id_personal_id ,
                    'usuario': usuario_list[0]['usuario'],
                    'estado': usuario_list[0]['estado'],
                    'nivel': usuario_list[0]['id_tipo_usuario_id'],
                }
                data = {
                    'resultado': 'no_ok',
                    'mensaje': 'Usuario existente desea modificarlo ?',
                    'usuario_list': usuario_list,
                }
            else:
                # ASIGNAMOS DATOS AL OBJETO
                usuario = Usuario(
                    id_personal_id=personal.id,
                    usuario=usuario,
                    clave=claveNueva,
                    id_tipo_usuario_id=id_tipo_usuario,
                    estado=estado,
                    fec_ingresa=fec_ingresa,
                    usuario_ing_id=usuario_ing_id,
                )
                # guardar datos
                usuario.save()
                # retornamos la respuesta
                data = {
                    'resultado': 'ok_insert',
                    'mensaje': 'Datos Guardados Correctamente !!',
                }
            return HttpResponse(json.dumps(data), content_type="application/json")


def updateUsuario(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        # validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            nombre = request.POST.get('id_personal')
            id_usuario = request.POST.get('id_usuario')
            usuario = request.POST.get('usuario')
            estado = request.POST.get('estado')

            fec_modifica = request.POST.get('datetime')
            usuario_mod_id = request.session['idUsuario']
            personal = Personal.objects.get(nombre=nombre)

            # clave foranea
            nivel = TipoUsuario.objects.get(pk=id_tipo_usuario)

            if request.POST.get('clave'):
                clave = request.POST.get('clave')
                claveNueva = md5.new(clave).hexdigest()
                # update con clave
                Usuario.objects.filter(pk=id_usuario).update(
                    id_personal_id=personal.id,
                    usuario=usuario,
                    clave=claveNueva,
                    id_tipo_usuario_id=nivel,
                    estado=estado,
                    fec_modifica=fec_modifica,
                    usuario_mod_id=usuario_mod_id,
                )
            else:
                # update sin clave
                Usuario.objects.filter(pk=id_usuario).update(
                    id_personal_id=personal.id,
                    usuario=usuario,
                    id_tipo_usuario_id=nivel,
                    estado=estado,
                    fec_modifica=fec_modifica,
                    usuario_mod_id=usuario_mod_id,
                )

            data = {
                'resultado': 'ok_update',
                'mensaje': 'Datos Actualizados Correctamente !!',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")



def excelUsuarios(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        #Consultar las Bodegas
        #usuarios = Usuario.objects.all().order_by('nombre')
        cursor = connection.cursor()
        query = " SELECT seguridad_usuario.usuario,  seguridad_usuario.estado, seguridad_usuario.id_tipo_usuario_id, seguridad_usuario.id, "
        query = query + " seguridad_usuario.id_personal_id, "
        query = query + " personal_personal.nombre AS nombre_personal, personal_personal.telefono, personal_personal.celular, personal_personal.correo, seguridad_usuario.id_tipo_usuario_id  "
        #se agrega para mostrar proceso
        #query = query + " areas_procesos.proceso "
        query = query + " FROM seguridad_usuario "
        query = query + " LEFT JOIN personal_personal ON seguridad_usuario.id_personal_id = personal_personal.id ORDER BY nombre_personal "
        #se agrega para mostrar proceso
        #query = query + " LEFT JOIN areas_area_proceso ON areas_area_proceso.personal_id = personal_personal.id "
        #query = query + " LEFT JOIN areas_procesos ON areas_procesos.id = areas_area_proceso.proceso_id "

        cursor.execute(query)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        usuarios_list = list(rows)

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Usuarios"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:I2')
        ws['B2'] = 'Reporte General de Usuarios'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True)


        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True)

        ws['D3'] = 'Teléfono'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True)

        ws['E3'] = 'Celular'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True)

        ws['F3'] = 'Correo'
        ws.cell(row=3, column=6).border = thin_border
        ws.cell(row=3, column=6).alignment = Alignment(horizontal='center')
        f3=ws['F3']
        f3.font = Font( bold=True)
        
        ws['G3'] = 'Usuario'
        ws.cell(row=3, column=7).border = thin_border
        ws.cell(row=3, column=7).alignment = Alignment(horizontal='center')
        g3=ws['G3']
        g3.font = Font( bold=True)
        
        ws['H3'] = 'Nivel'
        ws.cell(row=3, column=8).border = thin_border
        ws.cell(row=3, column=8).alignment = Alignment(horizontal='center')
        h3=ws['H3']
        h3.font = Font( bold=True)

        #ws['I3'] = 'Proceso'
        #ws.cell(row=3, column=9).border = thin_border
        #ws.cell(row=3, column=9).alignment = Alignment(horizontal='center')
        #i3=ws['I3']
        #i3.font = Font( bold=True)


        cont = 4
        indice = 1
        for usuario in usuarios_list:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = usuario['nombre_personal']
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            if usuario['telefono']:

                ws.cell(row=cont,column=4).value = usuario['telefono']
            else:
                ws.cell(row=cont,column=4).value = "--"
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            if usuario['celular']: 
                ws.cell(row=cont,column=5).value = usuario['celular']
            else:
              ws.cell(row=cont,column=5).value = "--"  
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=6).value = usuario['correo']
            ws.cell(row=cont,column=6).border = thin_border
            ws.cell(row=cont,column=6).alignment = Alignment(horizontal='center')

             #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=7).value = usuario['usuario']
            ws.cell(row=cont,column=7).border = thin_border
            ws.cell(row=cont,column=7).alignment = Alignment(horizontal='center')

            #asignando el cargo
            if usuario['id_tipo_usuario_id'] == 1:
                ws.cell(row=cont,column=8).value = "Administrador"

            if usuario['id_tipo_usuario_id'] == 2:
                ws.cell(row=cont,column=8).value = "Auditor"

            if usuario['id_tipo_usuario_id'] == 3 :
                ws.cell(row=cont,column=8).value = "Colaborador"

            if usuario['id_tipo_usuario_id'] == 4 :
                ws.cell(row=cont,column=8).value = "Director de Area"

            if usuario['id_tipo_usuario_id'] == 5:
                    ws.cell(row=cont, column=8).value = "Auditor Lider"

            if usuario['id_tipo_usuario_id'] == 6:
                ws.cell(row=cont, column=8).value = "Lider de Inocuidad"

            ws.cell(row=cont,column=8).border = thin_border
            ws.cell(row=cont,column=8).alignment = Alignment(horizontal='center')

            if usuario['estado'] == "0":

                #obtengo celdas de estado inactivo
                nombreCell=ws.cell(row=cont,column=3)
                telefonoCell=ws.cell(row=cont,column=4)
                celularCell=ws.cell(row=cont,column=5)
                correoCell=ws.cell(row=cont,column=6)
                usuarioCell=ws.cell(row=cont,column=7)
                nivelCell=ws.cell(row=cont,column=8)
                procesoCell=ws.cell(row=cont,column=9)

                #color rojo a inactivos
                nombreCell.font = Font(color=colors.RED)
                telefonoCell.font = Font(color=colors.RED)
                celularCell.font = Font(color=colors.RED)
                correoCell.font = Font(color=colors.RED)
                usuarioCell.font = Font(color=colors.RED)
                nivelCell.font= Font(color=colors.RED)
                procesoCell.font= Font(color=colors.RED)

            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Reporte_General_Usuarios.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 40.0
        ws.column_dimensions["D"].width = 25.0
        ws.column_dimensions["E"].width = 25.0
        ws.column_dimensions["F"].width = 30.0
        ws.column_dimensions["G"].width = 20.0
        ws.column_dimensions["H"].width = 20.0
        ws.column_dimensions["I"].width = 30.0
        wb.save(response)
        #retorna el archivo excel
        return response


def pdfUsuarios(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        
        cursor = connection.cursor()
        query = " SELECT seguridad_usuario.usuario,  seguridad_usuario.estado, seguridad_usuario.id_tipo_usuario_id, seguridad_usuario.id, "
        query = query + " seguridad_usuario.id_personal_id, "
        query = query + " personal_personal.nombre AS nombre_personal, personal_personal.telefono,  personal_personal.celular, personal_personal.correo, personal_personal.id_cargo_id, "
        query = query + " seguridad_tipousuario.tipo AS cargo "
        query = query + " FROM seguridad_usuario "
        query = query + " LEFT JOIN personal_personal ON seguridad_usuario.id_personal_id = personal_personal.id "
        query = query + " LEFT JOIN seguridad_tipousuario ON seguridad_usuario.id_tipo_usuario_id = seguridad_tipousuario.id "
        query = query + " ORDER BY cargo "

        cursor.execute(query)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        usuarios_list = list(rows)

        context = {
            'usuarios': usuarios_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfUsuarios.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Listado_Usuarios_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response



def dictfetchall(cursor):
    "Retorna data como diccionario y con keys"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]

        

def insertPermisos(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_usuario=request.POST.get('id_usuario')
            has_personal=request.POST.get('personal')
            has_usuario=request.POST.get('usuario')
            has_area=request.POST.get('area')
            has_norma=request.POST.get('norma')
            has_auditoria=request.POST.get('auditoria')
            has_documental=request.POST.get('documental')
            has_sac=request.POST.get('sac')

            Usuario.objects.filter(pk=id_usuario).update(
                has_personal=has_personal,
                has_usuario=has_usuario,
                has_area=has_area,
                has_norma=has_norma,
                has_auditoria=has_auditoria,
                has_documental=has_documental,
                has_sac=has_sac,
            )

            data = {
                'resultado':'ok_insert',
                'mensaje': 'Datos Grabados Correctamente !!',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")



def selectPermisos(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_usuario = id_usuario=request.POST.get('id_usuario')

            permisos=Usuario.objects.values('has_area', 'has_auditoria', 'has_documental', 'has_norma', 'has_personal', 'has_usuario', 'has_sac').all().filter(pk=id_usuario)
            permisos_list=list(permisos)

            data = {
                'resultado':'ok_select',
                'permisos':permisos_list,
            }
            return HttpResponse(json.dumps(data), content_type="application/json")


"""
NOTIFICACIONES
"""

def notificacion_colaborador(f):
    def wrap(request, *args, **kwargs):

            if 'idUsuario' not in request.session:
                return redirect('login')

            #eliminar notificaciones
            if 'notificacionSeguimiento' in request.session:
                del request.session['notificacionSeguimiento']

            if 'notificacionPlanAccion' in request.session:
                del request.session['notificacionPlanAccion']

            if 'notificacionDocumentosPendientes' in request.session:
                del request.session['notificacionDocumentosPendientes']

            # notificacion SAC COLABORADOR SEGUIMIENTO
            seguimiento_list = list(PlanAccion.objects.values('estado_seguimiento').all().filter(
            Q(estado_seguimiento=1) | Q(estado_seguimiento=4), responsable_seguimiento_id=request.session['idUsuario']))
            request.session['notificacionSeguimiento'] = seguimiento_list

            # NOTIFICACION PLAN ACCION
            cursor = connection.cursor()
            query = " SELECT COUNT(sac_planaccion.detalle_plan_accion) AS total FROM sac_planaccion WHERE sac_planaccion.responsable_plan_accion_id=%s AND sac_planaccion.justificacion_plan IS  null "
            params = [request.session['idUsuario']]
            cursor.execute(query, params)

            rows = dictfetchall(cursor)
            for row in rows:
                if row['total'] == 0:
                    if 'notificacionPlanAccion' in request.session:
                        del request.session['notificacionPlanAccion']
                else:
                    request.session['notificacionPlanAccion'] = rows
                    break

            # notificacion PENDIENTE SUBIR ARCHIVO

            cursor = connection.cursor()
            query = " SELECT DISTINCT areas_procesos.proceso AS proceso "
            query = query + " FROM documentos_documento "
            query = query + " LEFT JOIN areas_procesos ON documentos_documento.procedimiento_id = areas_procesos.id "
            query = query + " WHERE (documentos_documento.estado = 0 OR documentos_documento.estado = 2) "
            query = query + " AND documentos_documento.subido_por_id =%s "
            params = [request.session['idUsuario']]
            cursor.execute(query, params)

            rows = dictfetchall(cursor)
            documentos_pendientes_list = list(rows)
            total_elem = len(documentos_pendientes_list)
            if total_elem == 0:
                if 'notificacionDocumentosPendientes' in request.session:
                    del request.session['notificacionDocumentosPendientes']
            else:

                request.session['notificacionDocumentosPendientes'] = documentos_pendientes_list

            return f(request, *args, **kwargs)
    wrap.__doc__=f.__doc__
    wrap.__name__=f.__name__
    return wrap



def notificacion_director(f):
    def wrap(request, *args, **kwargs):

        if 'idUsuario' not in request.session:
            return  redirect('login')

        #eliminar notificaciones
        if 'notificacionJefe' in request.session:
            del request.session['notificacionJefe']

        if 'notificacionDocPorRevisar' in request.session:
            del request.session['notificacionDocPorRevisar']

        #NOTIFICACION SAC PENDIENTE
        sac_list = list(Sac.objects.values('estado_cabecera').all().filter(sac_jefe=0, responsable_id=request.session['idUsuario']))
        total_elment = len(sac_list)
        if total_elment == 0:
            if 'notificacionJefe' in request.session:
                del request.session['notificacionJefe']
        else:
            request.session['notificacionJefe'] = sac_list


        #NOTIFICACION POE PENDIENTE
        cursor = connection.cursor()
        query = " SELECT DISTINCT areas_procesos.proceso AS proceso, documentos_documento.nombre "
        query = query + " FROM documentos_documento "
        query = query + " LEFT JOIN areas_procesos ON documentos_documento.procedimiento_id = areas_procesos.id "
        query = query + " LEFT JOIN documentos_revision ON documentos_documento.id = documentos_revision.documento_id "
        #query = query + " WHERE documentos_documento.estado = 1  OR  documentos_documento.estado = 3 "
        query = query + " WHERE documentos_documento.estado = 1 "
        query = query + " AND documentos_revision.director_id =%s "

        params = [request.session['idUsuario']]
        cursor.execute(query, params)

        rows = dictfetchall(cursor)
        documentos_pendientes_list = list(rows)
        # total records de la consulta
        total_elementos = len(documentos_pendientes_list)
        # si la consulta trae al menos 1 elemento se realiza la asignacion
        if total_elementos >= 1:
            request.session['notificacionDocPorRevisar'] = documentos_pendientes_list


        return f(request, *args, **kwargs)

    wrap.__doc__ = f.__doc__
    wrap.__name__ = f.__name__
    return wrap


def notificacion_lider_norma(f):
    def wrap(request, *args, **kwargs):


        if 'idUsuario' not in request.session:
            return  redirect('login')

        #eliminar notificaciones
        if 'notificacionLiderNorma' in request.session:
            del request.session['notificacionLiderNorma']

        #NOTIFICACIONES POES PENDIENTES
        cursor = connection.cursor()
        query = " SELECT DISTINCT areas_procesos.proceso AS proceso, documentos_documento.nombre "
        query = query + " FROM documentos_documento "
        query = query + " LEFT JOIN areas_procesos ON documentos_documento.procedimiento_id = areas_procesos.id "
        query = query + " LEFT JOIN documentos_revision ON documentos_documento.id = documentos_revision.documento_id "
        query = query + " WHERE documentos_documento.estado = 1 OR documentos_documento.estado = 2 OR documentos_documento.estado = 3 "
        query = query + " AND documentos_revision.lider_id =%s "
        params = [request.session['idUsuario']]

        cursor.execute(query, params)
        rows = dictfetchall(cursor)

        documentos_pendientes_list = list(rows)
        # total records de la consulta
        total_elementos = len(documentos_pendientes_list)

        # si la consulta trae al menos 1 elemento se realiza la asignacion
        if total_elementos >= 1:
            request.session['notificacionLiderNorma'] = documentos_pendientes_list

        return f(request, *args, **kwargs)

    wrap.__doc__ = f.__doc__
    wrap.__name__ = f.__name__
    return wrap



def notificacion_auditor(f):
    def wrap(request, *args, **kwargs):

        if 'idUsuario' not in request.session:
            return  redirect('login')

        if 'notificacionAuditor' in request.session:
            del request.session['notificacionAuditor']

        #NOTIFICACIONES SAC AUDITOR
        sac_list = list(Sac.objects.values('estado_cabecera').all().filter(estado_cabecera=2))
        request.session['notificacionAuditor'] = sac_list

        return f(request, *args, **kwargs)

    wrap.__doc__ = f.__doc__
    wrap.__name__ = f.__name__
    return wrap


def ConvertDocToPDF(request):
    """
    Convierte un documento a pdf
    :param request: path
    :return: pdf convertido
    """
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        path = request.POST.get('path_documento')
        #from django.core.management import call_command
        #call_command('unoconv - f pdf '+path)
        import os
        import subprocess

        #eliminar archivo anterior
        try:
            os.remove('media/Convertidos/lectura.pdf')
        except OSError:
            pass

        #convirtiendo documento
        try:
            subprocess.check_call(
                ['/usr/bin/python2.7', '/usr/bin/unoconv', '-f', 'pdf', '-o', 'media/Convertidos/lectura.pdf', '-d', 'document',path])
        except subprocess.CalledProcessError as e:
            print('CalledProcessError', e)

        data = {
            'resultado' : 'ok_convert'
        }
        return JsonResponse(data, safe=False)



def GridDocumentosDisponibles(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:


        cursor = connection.cursor()

        query = " SELECT documentos_documento.id, documentos_documento.nombre, documentos_documento.version, "
        query = query + " documentos_documento.descripcion, documentos_documento.estado, areas_areas.area AS proceso, "
        query = query + " areas_procesos.proceso AS procedimiento "
        query = query + " FROM documentos_documento "
        query = query + " LEFT JOIN areas_areas ON documentos_documento.proceso_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON documentos_documento.procedimiento_id = areas_procesos.id "
        query = query + " WHERE documentos_documento.estado = 4 "

        cursor.execute(query)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        documentos_list = list(rows)

        if documentos_list:
            #RETORNA
            return JsonResponse(documentos_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
                'resultado':'no_ok',
                'mensaje': 'no existen Documentos',
            }
            return JsonResponse(json, safe=False)

@notificacion_colaborador
@notificacion_director
def Unificado(request):
        context={
        }
        return render(request, 'user/principal.html', context)



def InsertarUsuarioMultiple(request):

    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:

        # validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            nombre = request.POST.get('id_personal')
            usuario = request.POST.get('usuario')
            clave = request.POST.get('clave')
            claveNueva = md5.new(clave).hexdigest()
            estado = request.POST.get('estado')
            fec_ingresa = request.POST.get('datetime')
            usuario_ing_id = request.session['idUsuario']
            if request.POST.get('director'):
                director = request.POST.get('director')
            else:
                director = None

            if request.POST.get('colaborador'):
                colaborador = request.POST.get('colaborador')
            else:
                colaborador = None

            if request.POST.get('lider_norma'):
                lider_norma = request.POST.get('lider_norma')
            else:
                lider_norma = None

            if request.POST.get('auditor'):
                auditor = request.POST.get('auditor')
            else:
                auditor = None

            if request.POST.get('auditor_lider'):
                auditor_lider = request.POST.get('auditor_lider')
            else:
                auditor_lider = None

            if request.POST.get('permiso_lectura'):
                permiso_lectura = request.POST.get('permiso_lectura')
            else:
                permiso_lectura = None

            if request.POST.get('permiso_escritura'):
                permiso_escritura = request.POST.get('permiso_escritura')
            else:
                permiso_escritura = None


            try:
                personal = Personal.objects.get(nombre=nombre)
            except Personal.DoesNotExist:
                data = {
                    'resultado': 'no_ok_personal',
                    'mensaje': 'No existe el Colaborador !!',
                }
                return HttpResponse(json.dumps(data), content_type="application/json")


            cursor = connection.cursor()
            query = " SELECT * "
            query = query + " FROM seguridad_usuario "
            query = query + " WHERE seguridad_usuario.usuario= %s "
            params = [usuario]

            cursor.execute(query, params)
            rows = dictfetchall(cursor)
            usuario_list = list(rows)
            total_elementos = len(usuario_list)

            if total_elementos >= 1:
                # consultar bodega por nombre
                usuario_list = {
                    'idUsuario': usuario_list[0]['id'],
                    'id_personal': personal.nombre,
                    # 'id_personal':usuario.id_personal_id ,
                    'usuario': usuario_list[0]['usuario'],
                    'estado': usuario_list[0]['estado'],
                    'nivel': usuario_list[0]['id_tipo_usuario_id'],
                }
                data = {
                    'resultado': 'no_ok',
                    'mensaje': 'Usuario existente desea modificarlo ?'
                    #'usuario_list': usuario_list,
                }
                return HttpResponse(json.dumps(data), content_type="application/json")
            else:
                # ASIGNAMOS DATOS AL OBJETO
                usuario = Usuario(
                    id_personal_id=personal.id,
                    usuario=usuario,
                    clave=claveNueva,
                    id_tipo_usuario_id=None,
                    estado=estado,
                    fec_ingresa=fec_ingresa,
                    usuario_ing_id=usuario_ing_id,
                    director = director,
                    colaborador = colaborador,
                    lider_norma = lider_norma,
                    auditor = auditor,
                    auditor_lider = auditor_lider,
                    permiso_lectura = permiso_lectura,
                    permiso_escritura = permiso_escritura
                )
                # guardar datos
                usuario.save()
                # retornamos la respuesta
                data = {
                    'resultado': 'ok_insert',
                    'mensaje': 'Datos Guardados Correctamente !!',
                }
                return HttpResponse(json.dumps(data), content_type="application/json")



def UpdateUsuarioMultiple(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        # validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            nombre = request.POST.get('id_personal')
            id_usuario = request.POST.get('id_usuario')
            usuario = request.POST.get('usuario')
            estado = request.POST.get('estado')

            if request.POST.get('director'):
                director = request.POST.get('director')
            else:
                director = None

            if request.POST.get('colaborador'):
                colaborador = request.POST.get('colaborador')
            else:
                colaborador = None

            if request.POST.get('lider_norma'):
                lider_norma = request.POST.get('lider_norma')
            else:
                lider_norma = None

            if request.POST.get('auditor'):
                auditor = request.POST.get('auditor')
            else:
                auditor = None

            if request.POST.get('auditor_lider'):
                auditor_lider = request.POST.get('auditor_lider')
            else:
                auditor_lider = None

            if request.POST.get('permiso_lectura'):
                permiso_lectura = request.POST.get('permiso_lectura')
            else:
                permiso_lectura = None

            if request.POST.get('permiso_escritura'):
                permiso_escritura = request.POST.get('permiso_escritura')
            else:
                permiso_escritura = None

            fec_modifica = request.POST.get('datetime')
            usuario_mod_id = request.session['idUsuario']
            personal = Personal.objects.get(nombre=nombre)


            if request.POST.get('clave'):

                clave = request.POST.get('clave')
                claveNueva = md5.new(clave).hexdigest()
                # update con clave
                Usuario.objects.filter(pk=id_usuario).update(
                    id_personal_id=personal.id,
                    usuario=usuario,
                    clave=claveNueva,
                    #id_tipo_usuario_id=None,
                    estado=estado,
                    director = director,
                    colaborador = colaborador,
                    lider_norma = lider_norma,
                    auditor = auditor,
                    auditor_lider = auditor_lider,
                    permiso_lectura = permiso_lectura,
                    permiso_escritura = permiso_escritura,
                    fec_modifica=fec_modifica,
                    usuario_mod_id=usuario_mod_id
                )
            else:
                # update sin clave
                Usuario.objects.filter(pk=id_usuario).update(
                    id_personal_id=personal.id,
                    usuario=usuario,
                    id_tipo_usuario_id=None,
                    estado=estado,
                    director=director,
                    colaborador=colaborador,
                    lider_norma=lider_norma,
                    auditor=auditor,
                    auditor_lider=auditor_lider,
                    permiso_lectura=permiso_lectura,
                    permiso_escritura=permiso_escritura,
                    fec_modifica=fec_modifica,
                    usuario_mod_id=usuario_mod_id,
                )

            data = {
                'resultado': 'ok_update',
                'mensaje': 'Datos Actualizados Correctamente !!',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")


@csrf_exempt
def SubirImagenCkeditor(request):
    if request.method == 'POST' and request.FILES:
        PATH_IMAGENES_SAC = "media/recursosSac/"
        #archivo
        myfile = request.FILES['upload']
        #parametros CKEDITOR
        CKEditorFuncNum = request.GET['CKEditorFuncNum']
        langCode = request.GET['langCode']
        CKEditor = request.GET['CKEditor']
        #SUBIR DOCUMENTO
        fs = FileSystemStorage(location=PATH_IMAGENES_SAC)
        filename = fs.save(myfile.name, myfile)
        uploaded_file_url = PATH_IMAGENES_SAC + filename

        #VALIDAR EXTENSION
        filename, file_extension = os.path.splitext(uploaded_file_url)
        if file_extension == '.jpg' or file_extension =='.png':
            message = ''
            URL_IMAGEN = "'"+uploaded_file_url+"'"
            HTML_CODIGO ="<script type='text/javascript'> window.parent.CKEDITOR.tools.callFunction("+CKEditorFuncNum+","+URL_IMAGEN+", 'Archivo subido !!')</script>"
            return HttpResponse(HTML_CODIGO)
        else:
            os.remove(uploaded_file_url)
            HTML_CODIGO = "<script type='text/javascript'>alert('Extensión no permitida !!');</script>"
            return HttpResponse(HTML_CODIGO)
    else:
        MENSAJE= "<script type='text/javascript'> alert('error - No se ha subido la imagen !!');</script>"
        return HttpResponse(MENSAJE)



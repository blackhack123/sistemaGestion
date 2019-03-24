# -*- coding: utf-8 -*-
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
import json
from django.contrib import messages
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#IMPORTAMOS BORDES 
from openpyxl.styles import Border, Side, Alignment, Color, Fill, colors, Font
from openpyxl.cell import Cell
#importamos render_to_pdf de vista principal
from principal.views import render_to_pdf
from django.core.files.storage import FileSystemStorage
#IMPORTAMOS LA CONEXION PARA QUERY PERSONALIZADO
from django.db import connection
#importamos el convert diccionario
from principal.views import dictfetchall
from personal.forms import CargoForm, PersonalForm, AreaForm
from personal.models import Cargo, Personal, AreaPersonal

#FUNCIONES/VISTAS

def personal(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        cursor = connection.cursor()
        query = "SELECT personal_personal.id as id_personal, personal_personal.nombre AS nombre_persona, personal_personal.telefono, "
        query = query + " personal_personal.celular, personal_personal.correo, personal_personal.estado, "
        query = query + " personal_personal.firma, personal_cargo.nombre AS cargo, personal_personal.id_cargo_id,  "
        query = query + " personal_areapersonal.nombre AS area "
        query = query + " FROM personal_personal "
        query = query + " LEFT JOIN personal_cargo ON personal_personal.id_cargo_id = personal_cargo.id "
        query = query + " LEFT JOIN personal_areapersonal ON personal_personal.id_areapersonal_id = personal_areapersonal.id "
        query = query + " ORDER BY nombre_persona "

        cursor.execute(query)
        rows = dictfetchall(cursor)
        personal_list = list(rows)



        context={
            'personal':personal_list,
            'personalForm':PersonalForm(),
            'cargos':Cargo.objects.values('id','nombre').all().filter(estado=1),
            'area':AreaPersonal.objects.values('id','nombre').all().filter(estado=1),
        }

        return render(request, 'personal/personal.html', context)


def cargo(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:


        #variable de session para validar update/insert
        if 'insert_area' in request.session:
            insert_area = 'ok_insert'
            del request.session['insert_area']

        elif 'update_area' in request.session:
            insert_area = 'ok_insert'
            del request.session['update_area']
        else:
            insert_area = None

        context={
            'cargoForm':CargoForm(),
            'cargos':Cargo.objects.values().all(),
            'areaForm':AreaForm(),
            'area':AreaPersonal.objects.values().all().order_by('nombre'),
            'insert_area':insert_area
        }

        return render(request, 'personal/cargo.html', context)
        


def insertCargo(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        form = CargoForm(request.POST)
        
        cargo = Cargo.objects.filter(nombre=form['nombre'].value()).count()
        if cargo == 0:
            cargo=Cargo(
                nombre=form['nombre'].value(),
                estado=form['estado'].value(),
                descripcion=form['descripcion'].value(),
            )
            cargo.save()



            messages.add_message(request, messages.SUCCESS, 'Registro Creado Existosamente !!')
            return redirect('cargo')
        else:
            messages.add_message(request, messages.WARNING, 'El registro ya existe.. no puede ser creado !!')
            return redirect('cargo')


def selectCargo(request):
        
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_cargo = request.POST.get('id_cargo')

            cargo = Cargo.objects.get(pk=id_cargo)

            cargo_list ={
                'id':cargo.id,
                'nombre':cargo.nombre ,
                'descripcion':cargo.descripcion,
                'estado':cargo.estado,
            }
    
            data = {
            'resultado':'ok_select',
            'cargo_list':cargo_list,
            }
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})


def updateCargo(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        form = CargoForm(request.POST)
        
        Cargo.objects.filter(pk=form['pk'].value()).update(
            nombre=form['nombre'].value(),
            estado=form['estado'].value(),
            descripcion=form['descripcion'].value(),
        )
        messages.add_message(request, messages.SUCCESS, 'Actualizado Existosamente !!')
        return redirect('cargo')

def excelCargos(request):


     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        #Consultar las Bodegas
        cargos = Cargo.objects.all().order_by('nombre')

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Cargos"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:D2')
        ws['B2'] = 'Listado de Cargos'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Cargo'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Descripción'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)



        cont = 4
        indice = 1
        for cargo in cargos:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = cargo.nombre
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=4).value = cargo.descripcion
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')



            if cargo.estado == "0":

                #obtengo celdas de estado inactivo
                nombreCell=ws.cell(row=cont,column=3)
                descripcionCell=ws.cell(row=cont,column=4)

                #color rojo a inactivos
                nombreCell.font = Font(color=colors.RED)
                descripcionCell.font = Font(color=colors.RED)


            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Listado_Cargos.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 40.0
        ws.column_dimensions["E"].width = 40.0
        wb.save(response)
        #retorna el archivo excel
        return response

def pdfCargos(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        context = {
            'cargos': Cargo.objects.all(),
        }

        pdf = render_to_pdf('reportes/pdf/pdfCargos.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Listado_Cargos_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response



def selectPersonal(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_personal = request.POST.get('id_personal')

            personal = Personal.objects.get(pk=id_personal)

            personal_list ={
                'id':personal.id,
                'nombre':personal.nombre ,
                'telefono':personal.telefono,
                'celular':personal.celular,
                'correo':personal.correo,
                'estado':personal.estado,
                'id_cargo':personal.id_cargo_id,
                'id_area':personal.id_areapersonal_id,
            }
    
            data = {
            'resultado':'ok_select',
            'personal_list':personal_list,
            }
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})


def insertPersonal(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        form = PersonalForm(request.POST)

        personal = Personal.objects.filter(nombre=form['nombre'].value()).count()

        if personal == 0:
            if request.method == 'POST' and request.FILES:
                myfile = request.FILES['firma']
                destino_archivo = 'media/firmas'
                fs = FileSystemStorage(location=destino_archivo)
                filename = fs.save(myfile.name, myfile)
                uploaded_file_url = 'media/firmas/'+filename

                personal=Personal(
                    nombre=form['nombre'].value(),
                    telefono=form['telefono'].value(),
                    celular=form['celular'].value(),
                    correo=form['correo'].value(),
                    #estado=form['estado'].value(),
                    id_cargo_id=form['cargo'].value(),
                    id_areapersonal_id=form['area'].value(),
                    firma = uploaded_file_url.encode('utf-8')
                )
                personal.save()
            else:
                personal=Personal(
                    nombre=form['nombre'].value(),
                    telefono=form['telefono'].value(),
                    celular=form['celular'].value(),
                    correo=form['correo'].value(),
                    #estado=form['estado'].value(),
                    id_cargo_id=form['cargo'].value(),
                    id_areapersonal_id=form['area'].value()
                )
                personal.save()

            messages.add_message(request, messages.SUCCESS, 'Registro Creado Existosamente !!')
            return redirect('personal')
        else:
            messages.add_message(request, messages.WARNING, 'El registro ya existe.. no puede ser creado !!')
            return redirect('personal')

def updatePersonal(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        form = PersonalForm(request.POST)

        if request.method == 'POST' and request.FILES:

            #guardar firma
            myfile = request.FILES['firma']
            destino_archivo = 'media/firmas'
            fs = FileSystemStorage(location=destino_archivo)
            filename = fs.save(myfile.name.encode('utf-8'), myfile)
            uploaded_file_url = 'media/firmas/' + filename

            Personal.objects.filter(pk=form['pk'].value()).update(
                nombre=form['nombre'].value(),
                telefono=form['telefono'].value(),
                celular=form['celular'].value(),
                correo=form['correo'].value(),
                #estado=form['estado'].value(),
                id_cargo_id=form['cargo'].value(),
                id_areapersonal_id=form['area'].value(),
                firma=uploaded_file_url.encode('utf-8')
            )
        else:
            Personal.objects.filter(pk=form['pk'].value()).update(
                nombre=form['nombre'].value(),
                telefono=form['telefono'].value(),
                celular=form['celular'].value(),
                correo=form['correo'].value(),
                #estado=form['estado'].value(),
                id_cargo_id=form['cargo'].value(),
                id_areapersonal_id=form['area'].value()
            )

        messages.add_message(request, messages.SUCCESS, 'Registro Actualizado Existosamente !!')
        return redirect('personal')



def excelPersonal(request):


     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        #Consultar las Bodegas
        cursor = connection.cursor()
        query = "SELECT personal_personal.id as id_personal, personal_personal.nombre AS nombre_persona, "
        query = query + " personal_personal.telefono, personal_personal.celular,  personal_personal.correo, personal_personal.estado, "
        query = query + " personal_cargo.nombre AS cargo, personal_areapersonal.nombre AS area, "
         #se agrega para mostrar los procesos vinculados
        query = query + " areas_procesos.proceso "
        query = query + " FROM personal_personal "
        query = query + " LEFT JOIN personal_cargo ON personal_personal.id_cargo_id = personal_cargo.id "
        #se agrega para mostrar los procesos vinculados
        query = query + " LEFT JOIN areas_area_proceso ON areas_area_proceso.personal_id = personal_personal.id "
        query = query + " LEFT JOIN areas_procesos ON areas_procesos.id = areas_area_proceso.proceso_id "
        query = query + " LEFT JOIN personal_areapersonal ON personal_areapersonal.id = personal_personal.id_areapersonal_id "
        query = query + " ORDER BY cargo "
        
        cursor.execute(query)
        rows = dictfetchall(cursor)
        personal_list = list(rows)

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Personal"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:H2')
        ws['B2'] = 'Listado de Personal'
        ws['B2'].alignment = Alignment(horizontal='center')
        #ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Nombre'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Área'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Cargo'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['F3'] = 'Proceso'
        ws.cell(row=3, column=6).border = thin_border
        ws.cell(row=3, column=6).alignment = Alignment(horizontal='center')
        f3=ws['F3']
        f3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['G3'] = 'Teléfono'
        ws.cell(row=3, column=7).border = thin_border
        ws.cell(row=3, column=7).alignment = Alignment(horizontal='center')
        g3=ws['G3']
        g3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['H3'] = 'Celular'
        ws.cell(row=3, column=8).border = thin_border
        ws.cell(row=3, column=8).alignment = Alignment(horizontal='center')
        h3=ws['H3']
        h3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['I3'] = 'Correo'
        ws.cell(row=3, column=9).border = thin_border
        ws.cell(row=3, column=9).alignment = Alignment(horizontal='center')
        i3=ws['I3']
        i3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        cont = 4
        indice = 1
        for personal in personal_list:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = personal['nombre_persona']
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            if personal['area']:
                ws.cell(row=cont,column=4).value = personal['area']
            else:
                ws.cell(row=cont,column=4).value = '--'            
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=5).value = personal['cargo']
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            if personal['proceso']:
                ws.cell(row=cont,column=6).value = personal['proceso']
            else:
                ws.cell(row=cont,column=6).value = '--'
            ws.cell(row=cont,column=6).border = thin_border
            ws.cell(row=cont,column=6).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            if personal['telefono']:
                ws.cell(row=cont,column=7).value = personal['telefono']
            else:
                ws.cell(row=cont,column=7).value = '--'
            ws.cell(row=cont,column=7).border = thin_border
            ws.cell(row=cont,column=7).alignment = Alignment(horizontal='center')


            #agregamos la data, borde, alineacion
            if personal['celular']:
                ws.cell(row=cont,column=8).value = personal['celular']
            else:
                ws.cell(row=cont,column=8).value = '--'
            ws.cell(row=cont,column=8).border = thin_border
            ws.cell(row=cont,column=8).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=9).value = personal['correo']
            ws.cell(row=cont,column=9).border = thin_border
            ws.cell(row=cont,column=9).alignment = Alignment(horizontal='center')



            if personal['estado'] == "0":

                #obtengo celdas de estado inactivo
                nombreCell=ws.cell(row=cont,column=3)
                telefonoCell=ws.cell(row=cont,column=4)
                correoCell=ws.cell(row=cont,column=5)
                cargoCell=ws.cell(row=cont,column=6)
                celularCell=ws.cell(row=cont,column=7)
                areaCell=ws.cell(row=cont,column=8)
                procesoCell=ws.cell(row=cont,column=9)

                #color rojo a inactivos
                nombreCell.font = Font(color=colors.RED)
                telefonoCell.font = Font(color=colors.RED)
                correoCell.font = Font(color=colors.RED)
                cargoCell.font = Font(color=colors.RED)
                celularCell.font = Font(color=colors.RED)
                areaCell.font = Font(color=colors.RED)
                procesoCell.font = Font(color=colors.RED)



            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Listado_Personal.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 20.0
        ws.column_dimensions["E"].width = 20.0
        ws.column_dimensions["F"].width = 20.0
        ws.column_dimensions["G"].width = 20.0
        ws.column_dimensions["H"].width = 20.0
        ws.column_dimensions["I"].width = 40.0
        wb.save(response)
        #retorna el archivo excel
        return response



def pdfPersonal(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        #Consultar las Bodegas
        cursor = connection.cursor()
        query = "SELECT personal_personal.id as id_personal, personal_personal.nombre AS nombre_persona, "
        query = query + " personal_personal.telefono, personal_personal.celular,  personal_personal.correo, personal_personal.estado, "
        query = query + " personal_cargo.nombre AS cargo, personal_areapersonal.nombre AS area, "
         #se agrega para mostrar los procesos vinculados
        query = query + " areas_procesos.proceso "
        query = query + " FROM personal_personal "
        query = query + " LEFT JOIN personal_cargo ON personal_personal.id_cargo_id = personal_cargo.id "
        #se agrega para mostrar los procesos vinculados
        query = query + " LEFT JOIN areas_area_proceso ON areas_area_proceso.personal_id = personal_personal.id "
        query = query + " LEFT JOIN areas_procesos ON areas_procesos.id = areas_area_proceso.proceso_id "
        query = query + " LEFT JOIN personal_areapersonal ON personal_areapersonal.id = personal_personal.id_areapersonal_id "
        query = query + " ORDER BY cargo "
        
        cursor.execute(query)
        rows = dictfetchall(cursor)
        personal_list = list(rows)

        

        context = {
            'personal_list': personal_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfPersonal.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Listado_de_Personal_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response


def perPorCargo(request):
    
    """
    Retorna listado del personal por cargo->id
    """
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():
            
            id_cargo = request.POST.get('id_cargo')
            
            personal = Personal.objects.all().filter(estado=1, id_cargo_id=id_cargo)

            cursor = connection.cursor()
            query = " SELECT * "
            query = query + " FROM personal_personal "
            query = query + " WHERE personal_personal.id_cargo_id =%s "
            query = query + " AND personal_personal.estado=1 "
            params=[id_cargo]

            cursor.execute(query, params)
            rows=dictfetchall(cursor)
           
            if rows:
                data = {
                    'resultado':'ok_select',
                    'personal_list': rows,
                }
            else:
                data = {
                    'resultado':'no_ok',
                }
          
            return JsonResponse(data)
        else:

            return render(request, 'seguridad/login.html', {})




def listadoPersonal(request):

 #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:


        cursor = connection.cursor()
        query = "SELECT personal_personal.id, personal_personal.nombre  FROM personal_personal WHERE estado = 1 "
        
        cursor.execute(query)
        rows=dictfetchall(cursor)



        return JsonResponse(rows, safe=False)

# -*- coding: utf-8 -*-
from django.contrib import messages
from django.shortcuts import render,redirect    
from django.http import HttpResponse, JsonResponse
import json
from django.core import serializers
#IMPORTAMOS LA CONEXION PARA QUERY PERSONALIZADO
from django.db import connection
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#IMPORTAMOS BORDES 
from openpyxl.styles import Border, Side, Alignment, Color, Fill, colors, Font
from openpyxl.cell import Cell
#importamos render_to_pdf de vista principal
from principal.views import render_to_pdf
#importamos el convert diccionario
from principal.views import dictfetchall
#importamos model Usuarios
from seguridad.models import Usuario
#importamos model Normas
from normas.models import Normas
#importamos el area
from areas.models import Areas
#importamos los procesos
from areas.models import Procesos
#importamos Auditores
from auditoria.models import Auditores,Auditorias
#importamos los cargos
from personal.models import Cargo, Personal
#IMPORTAMOS LIB DE EMIAL
#from django.core.mail import EmailMessage

# Create your views here.

def auditoria(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        context={
            'cargos':Cargo.objects.all().values('id', 'nombre').filter(estado=1),
            'normas':Normas.objects.all().values('id', 'nombre', 'estado').filter(estado=1),
        }
        return render(request, 'admin/auditoria.html', context)


def gridProDesignados(request):
        
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

            id_norma = request.GET['id_norma']

            cursor = connection.cursor()
            query = " SELECT normas_procesoclausula.id, areas_areas.area, areas_procesos.proceso, "
            query = query + " normas_procesoclausula.id_proceso_id,  normas_procesoclausula.id_area_id, normas_procesoclausula.id_norma_id, "
            query = query + "  normas_clausulas.clausula, normas_clausulas.id AS id_clausula "
            query = query + " FROM normas_procesoclausula "
            query = query + " LEFT JOIN areas_areas ON normas_procesoclausula.id_area_id = areas_areas.id "
            query = query + " LEFT JOIN areas_procesos ON normas_procesoclausula.id_proceso_id = areas_procesos.id "
            query = query + " LEFT JOIN normas_clausulas ON normas_procesoclausula.id_clausula_id = normas_clausulas.id "
            query = query + " WHERE normas_procesoclausula.id_norma_id = %s "
            #query = query + " GROUP BY normas_procesoclausula.id_proceso_id, normas_procesoclausula.id_clausula_id "

            params=str(id_norma)
            cursor.execute(query, params)
            rows = dictfetchall(cursor)

            procesos_list=list(rows)

            if procesos_list:

                #RETORNA  
                return JsonResponse(procesos_list, safe=False)

            else:
                #RETORNA MENSAJE DE NO EXISTEN
                json = {
                'resultado':'no_ok',
                'mensaje': 'no existen Procesos Designados',
                }
                return JsonResponse(json, safe=False)




def insertAuditor(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            data = json.loads(request.POST.get("data"))
            
            try:
                for auditor in data['gridData']:
                    auditorExistente=Auditores.objects.get(id_auditor_id=auditor['id_auditor'], id_norma_id=auditor['id_norma'], id_proceso_id=auditor['id_proceso'] )
                    personal=Personal.objects.get(pk=auditor['id_auditor'])
                    proceso=Procesos.objects.get(pk=auditor['id_proceso'])
                    auditor_list ={
                        'auditor':personal.nombre,
                        'proceso':proceso.proceso,
                    }
                    data = {
                        'resultado':'no_ok',
                        'auditor_list':auditor_list,
                    }
                    return HttpResponse(json.dumps(data), content_type="application/json")

            except Auditores.DoesNotExist :

                for auditor in data['gridData']:
                    auditor=Auditores(
                        #id_area_id=auditor['id_area'],
                        id_auditor_id=auditor['id_auditor'],
                        proceso_clausula_id=auditor['proceso_clausula']
                        #id_norma_id=auditor['id_norma'],
                        #id_proceso_id=auditor['id_proceso'],
                    )
                    auditor.save()
                data = {
                    'resultado':'ok_insert',
                    'mensaje': 'Datos guardados Correctamente  !',
                }

                return HttpResponse(json.dumps(data), content_type="application/json")



def gridAuditores(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:   

        #Consultar
        cursor = connection.cursor()
        """
        query = " SELECT DISTINCT "
        query = query + " normas_normas.nombre AS norma, "
        query = query + " personal_personal.nombre AS auditor, "
        query = query + " areas_areas.area, "
        query = query + " areas_procesos.proceso, "
        query = query + " auditoria_auditores.id, auditoria_auditores.id_auditor_id, auditoria_auditores.id_area_id, 	auditoria_auditores.id_norma_id, auditoria_auditores.id_proceso_id, "
        query = query + " normas_clausulas.clausula "
        query = query + " FROM auditoria_auditores "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditores.id_auditor_id = personal_personal.id"
        query = query + " LEFT JOIN normas_normas ON auditoria_auditores.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN areas_areas ON auditoria_auditores.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON auditoria_auditores.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN normas_clausulas ON normas_normas.id = normas_clausulas.id_norma_id "
        query = query + " GROUP BY norma, clausula, proceso "
        query = query + " ORDER BY areas_procesos.proceso "
        """
        query = " SELECT personal_personal.nombre AS auditor, normas_normas.nombre AS norma, normas_clausulas.clausula, "
        query = query + " areas_areas.area, areas_procesos.proceso, auditoria_auditores.id, auditoria_auditores.id_auditor_id, "
        query = query + " auditoria_auditores.id_area_id, auditoria_auditores.id_norma_id, auditoria_auditores.id_proceso_id "
        query = query + " FROM auditoria_auditores "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditores.id_auditor_id = personal_personal.id "
        query = query + " LEFT JOIN normas_procesoclausula ON auditoria_auditores.proceso_clausula_id = normas_procesoclausula.id "
        query = query + " LEFT JOIN normas_normas ON normas_procesoclausula.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN normas_clausulas ON normas_procesoclausula.id_clausula_id = normas_clausulas.id "
        query = query + " LEFT JOIN areas_areas ON normas_procesoclausula.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON normas_procesoclausula.id_proceso_id = areas_procesos.id  "
        query = query + " ORDER BY norma, clausula "


        cursor.execute(query)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        auditores_list = list(rows) 

        if auditores_list:
            #RETORNA  
            return JsonResponse(auditores_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Auditores Vinculados',
            }
            return JsonResponse(json, safe=False)



def excelAuditores(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        
        #Consultar
        cursor = connection.cursor()
        query = " SELECT " 
        query = query + " normas_normas.nombre AS norma, "
        query = query + " personal_personal.nombre AS auditor, "
        query = query + " areas_areas.area, "
        query = query + " areas_procesos.proceso, "
        query = query + " auditoria_auditores.id, auditoria_auditores.id_auditor_id, auditoria_auditores.id_area_id, 	auditoria_auditores.id_norma_id, auditoria_auditores.id_proceso_id"
        query = query + " FROM auditoria_auditores "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditores.id_auditor_id = personal_personal.id"
        query = query + " LEFT JOIN normas_normas ON auditoria_auditores.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN areas_areas ON auditoria_auditores.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON auditoria_auditores.id_proceso_id = areas_procesos.id "

        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        auditores_list = list(rows) 


        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Auditores"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:F2')
        ws['B2'] = 'Listado General de Auditores'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Auditor'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Norma'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Area'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['F3'] = 'Proceso'
        ws.cell(row=3, column=6).border = thin_border
        ws.cell(row=3, column=6).alignment = Alignment(horizontal='center')
        f3=ws['F3']
        f3.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        


        cont = 4
        indice = 1
        for auditor in auditores_list:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = auditor['auditor']
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=4).value = auditor['norma']
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=5).value = auditor['area']
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=6).value = auditor['proceso']
            ws.cell(row=cont,column=6).border = thin_border
            ws.cell(row=cont,column=6).alignment = Alignment(horizontal='center')



            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Listado_Auditores.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 40.0
        ws.column_dimensions["E"].width = 40.0
        ws.column_dimensions["F"].width = 40.0


        wb.save(response)
        #retorna el archivo excel
        return response


def pdfAuditores(request):
    
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #consultar
        cursor = connection.cursor()
        query = " SELECT " 
        query = query + " normas_normas.nombre AS norma, "
        query = query + " personal_personal.nombre AS auditor, "
        query = query + " areas_areas.area, "
        query = query + " areas_procesos.proceso, "
        query = query + " auditoria_auditores.id, auditoria_auditores.id_auditor_id, auditoria_auditores.id_area_id, 	auditoria_auditores.id_norma_id, auditoria_auditores.id_proceso_id"
        query = query + " FROM auditoria_auditores "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditores.id_auditor_id = personal_personal.id"
        query = query + " LEFT JOIN normas_normas ON auditoria_auditores.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN areas_areas ON auditoria_auditores.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON auditoria_auditores.id_proceso_id = areas_procesos.id "
        #query = query + " WHERE areas_procesos.id_area_id  IS NOT NULL "
        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        auditores_list = list(rows) 

        context = {
            'auditores_list': auditores_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfAuditores.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Listado_Auditores_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response




def removeAuditor(request):

        #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            id_registro = request.POST.get('id')
            
            Auditores.objects.filter(pk=id_registro ).delete()
            data = {
                'resultado':'ok_update',
                'mensaje': 'Auditor desvinculado exitosamente !!',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})


def gridAudtDesignados(request):

        #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:


        cursor = connection.cursor()
        """
        query = " SELECT " 
        query = query + " normas_normas.nombre AS norma, "
        query = query + " personal_personal.nombre AS auditor, "
        query = query + " areas_areas.area, "
        query = query + " areas_procesos.proceso, "
        query = query + " auditoria_auditores.id, auditoria_auditores.id_auditor_id, auditoria_auditores.id_area_id, 	auditoria_auditores.id_norma_id, auditoria_auditores.id_proceso_id, "
        query = query + " normas_clausulas.clausula "
        query = query + " FROM auditoria_auditores "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditores.id_auditor_id = personal_personal.id"
        query = query + " LEFT JOIN normas_normas ON auditoria_auditores.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN areas_areas ON auditoria_auditores.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON auditoria_auditores.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN normas_clausulas ON normas_normas.id = normas_clausulas.id_norma_id "
        query = query + " WHERE auditoria_auditores.id_norma_id = %s"
        query = query + " GROUP BY norma, proceso "
        query = query + " ORDER BY areas_procesos.proceso "
        """
        query = " SELECT personal_personal.id AS id_auditor_id, personal_personal.nombre AS auditor, normas_normas.nombre AS norma, "
        query = query + " normas_clausulas.id AS id_clausula, normas_clausulas.clausula, areas_areas.area, areas_procesos.proceso, "
        query = query + " normas_procesoclausula.id, normas_procesoclausula.id_area_id, normas_procesoclausula.id_norma_id, "
        query = query + " normas_procesoclausula.id_proceso_id "
        query = query + " FROM auditoria_auditores "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditores.id_auditor_id = personal_personal.id "
        query = query + " LEFT JOIN normas_procesoclausula ON auditoria_auditores.proceso_clausula_id = normas_procesoclausula.id "
        query = query + " LEFT JOIN normas_normas ON normas_procesoclausula.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN normas_clausulas ON normas_procesoclausula.id_clausula_id = normas_clausulas.id "
        query = query + " LEFT JOIN areas_areas ON normas_procesoclausula.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON normas_procesoclausula.id_proceso_id = areas_procesos.id "
        query = query + " WHERE normas_procesoclausula.id_norma_id =%s "
        params=str(request.GET['id_norma'])
  
        cursor.execute(query, params)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        auditores_list = list(rows) 

        if auditores_list:
            #RETORNA  
            return JsonResponse(auditores_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Auditores Vinculados',
            }
            return JsonResponse(json, safe=False)





def estabAuditoria(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        
        
        context={
      
            'normas':Normas.objects.all().values('id', 'nombre').filter(estado=1),
        }
        return render(request, 'admin/setAuditorias.html', context)


def setAuditoria(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            #data desde ajax(post)
            data = json.loads(request.POST.get("data"))
            
            
            for dataPost in data['gridData']:

                #auditoria = Auditorias.objects.filter(numero_auditoria=dataPost['numero'], id_proceso_id=dataPost['id_proceso']).count()
                #auditoria = Auditorias.objects.filter(numero_auditoria=dataPost['numero'], id_proceso_id=dataPost['id_proceso'], id_norma_id=dataPost['id_norma']).count()
                
                #if auditoria == 0:

                auditoria = Auditorias(

                    lugar=dataPost['lugar'],
                    fec_inicio=dataPost['fecha_inicio'],
                    fec_fin=dataPost['fecha_fin'],
                    hora_fin=dataPost['hora_fin'],
                    objetivo=dataPost['objetivo'],
                    id_area_id=dataPost['id_area'],
                    id_norma_id=dataPost['id_norma'],
                    hora_inicio=dataPost['hora_inicio'],
                    numero_auditoria=dataPost['numero'],
                    id_auditor_id=dataPost['id_auditor'],
                    id_proceso_id=dataPost['id_proceso'],
                    id_clausula_id = dataPost['id_clausula']
                )
                auditoria.save()

                #else:
                #data = {
                #    'resultado':'ok_select',
                #    'mensaje': 'Auditoría existente  !',
                #}
                #return HttpResponse(json.dumps(data), content_type="application/json")
            
                #enviar aviso auditoria establecida
                """
                destinatario="diegoavila@informaticos.com"
                destinatario2="juan.freire@rgmanagementcorp.com"
                email = EmailMessage(
                    'ROUTING GEAR: AUDITORIA ESTABLECIDA ', 
                    'FECHA INICIO: '+dataPost['fecha_inicio'] + "  Hora: "+dataPost['hora_inicio'] + " Auditoria N: "+dataPost['numero'] + " Lugar: "+dataPost['lugar'], 
                    to=[destinatario, destinatario2]
                )
                email.send()
                """
                #respuesta 
            
            data = {
                'resultado':'ok_insert',
                'mensaje': 'Datos guardados Correctamente  !',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")
            

        else:

            return render(request, 'seguridad/login.html', {})




def gridAuditorias(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

       

        cursor = connection.cursor()
        query = "SELECT auditoria_auditorias.id AS id_auditoria, auditoria_auditorias.lugar,  auditoria_auditorias.fec_inicio, auditoria_auditorias.fec_fin, "
        query = query + " auditoria_auditorias.objetivo, auditoria_auditorias.id_area_id, auditoria_auditorias.id_norma_id, "
        query = query + " auditoria_auditorias.hora_inicio, auditoria_auditorias.numero_auditoria, auditoria_auditorias.id_auditor_id, "
        query = query + " auditoria_auditorias.id_proceso_id, auditoria_auditorias.numero_auditoria, "
        query = query + " normas_normas.nombre AS nombre_norma, normas_clausulas.clausula AS clausula, "
        query = query + " areas_areas.area AS nombre_area, "
        query = query + " areas_procesos.proceso AS nombre_proceso, "
        query = query + " personal_personal.nombre AS nombre_auditor "
        query = query + " FROM auditoria_auditorias "
        query = query + " LEFT JOIN normas_normas on auditoria_auditorias.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN normas_clausulas ON auditoria_auditorias.id_clausula_id = normas_clausulas.id "
        query = query + " LEFT JOIN areas_areas on auditoria_auditorias.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos on auditoria_auditorias.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditorias.id_auditor_id = personal_personal.id"


        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        auditores_list = list(rows) 

        if auditores_list:
            #RETORNA  
            return JsonResponse(auditores_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Procesos',
            }
            return JsonResponse(json, safe=False)
        


def deleteAuditoria(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            numero_auditoria = request.POST.get("numero_auditoria")

            Auditorias.objects.filter(numero_auditoria=numero_auditoria).delete()

            data = {
                'resultado':'ok_delete',
                'mensaje': 'Auditoria removida Correctamente !!',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})


def getAuditoria(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            numero_auditoria = request.POST.get("numero_auditoria")

            #auditoria=Auditorias.objects.get(numero_auditoria=numero_auditoria)
            auditoria= Auditorias.objects.all().values().filter(numero_auditoria=numero_auditoria)[:1:1]

            auditoria_list=list(auditoria)
            #result=JsonResponse(auditoria_list, safe=False)
            
            data = {
                'resultado':'ok_select',
                'auditoria_list':auditoria_list,
            }
            return JsonResponse(data, safe=False)
            #return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})


def updateAudit(request):
    
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():


            numero = request.POST.get("numero")
            lugar = request.POST.get("lugar")
            fecha_inicio = request.POST.get("fecha_inicio")
            hora_inicio = request.POST.get("hora_inicio")
            fecha_fin = request.POST.get("fecha_fin")
            objetivo = request.POST.get("objetivo")

            #update sin clave
            Auditorias.objects.filter(numero_auditoria=numero).update(

                lugar=lugar,
                fec_inicio=fecha_inicio,
                hora_inicio=hora_inicio,
                fec_fin=fecha_fin, 
                objetivo=objetivo, 
                
            )

            data = {
                'resultado':'ok_update',
                'mensaje': 'Datos Actualizados Correctamente !!',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")


        else:

            return render(request, 'seguridad/login.html', {})



def excelAudits(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        
        #Consultar
        cursor = connection.cursor()
        """
        query = "SELECT auditoria_auditorias.id AS id_auditoria, auditoria_auditorias.lugar,  auditoria_auditorias.fec_inicio, auditoria_auditorias.fec_fin, "
        query = query + " auditoria_auditorias.objetivo, auditoria_auditorias.id_area_id, auditoria_auditorias.id_norma_id, "
        query = query + " auditoria_auditorias.hora_inicio, auditoria_auditorias.numero_auditoria, auditoria_auditorias.id_auditor_id, "
        query = query + " auditoria_auditorias.id_proceso_id, auditoria_auditorias.numero_auditoria, "
        query = query + " normas_normas.nombre AS nombre_norma, "
        query = query + " areas_areas.area AS nombre_area, "
        query = query + " areas_procesos.proceso AS nombre_proceso, "
        query = query + " areas_subproceso.nombre AS subproceso, "
        query = query + " personal_personal.nombre AS nombre_auditor "
        query = query + " FROM auditoria_auditorias "
        query = query + " LEFT JOIN normas_normas on auditoria_auditorias.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN areas_areas on auditoria_auditorias.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos on auditoria_auditorias.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN areas_subproceso ON auditoria_auditorias.id_sub_proceso_id = areas_subproceso.id_proceso_id "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditorias.id_auditor_id = personal_personal.id"
        query = query + " WHERE areas_subproceso.nombre IS NOT NULL"
        """
        query = "SELECT auditoria_auditorias.id AS id_auditoria, auditoria_auditorias.lugar,  auditoria_auditorias.fec_inicio, auditoria_auditorias.fec_fin, "
        query = query + " auditoria_auditorias.objetivo, auditoria_auditorias.id_area_id, auditoria_auditorias.id_norma_id, "
        query = query + " auditoria_auditorias.hora_inicio, auditoria_auditorias.numero_auditoria, auditoria_auditorias.id_auditor_id, "
        query = query + " auditoria_auditorias.id_proceso_id, auditoria_auditorias.numero_auditoria, "
        query = query + " normas_normas.nombre AS nombre_norma, "
        query = query + " areas_areas.area AS nombre_area, "
        query = query + " areas_procesos.proceso AS nombre_proceso, "
        query = query + " personal_personal.nombre AS nombre_auditor "
        query = query + " FROM auditoria_auditorias "
        query = query + " LEFT JOIN normas_normas on auditoria_auditorias.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN areas_areas on auditoria_auditorias.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos on auditoria_auditorias.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditorias.id_auditor_id = personal_personal.id"
        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        auditorias_list = list(rows) 

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Auditorias"

        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:K2')
        ws['B2'] = 'Reporte General de Auditorias'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        #Creamos los encabezados 
        ws['B3'] = 'AUDITORIA'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'NORMA'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'AREA'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'PROCESO'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        ws['F3'] = 'FEC. INICIO'
        ws.cell(row=3, column=6).border = thin_border
        ws.cell(row=3, column=6).alignment = Alignment(horizontal='center')
        f3=ws['F3']
        f3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['G3'] = 'HORA DE INICIO'
        ws.cell(row=3, column=7).border = thin_border
        ws.cell(row=3, column=7).alignment = Alignment(horizontal='center')
        g3=ws['G3']
        g3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        ws['H3'] = 'FECHA FINAL'
        ws.cell(row=3, column=8).border = thin_border
        ws.cell(row=3, column=8).alignment = Alignment(horizontal='center')
        h3=ws['H3']
        h3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['I3'] = 'AUDITOR'
        ws.cell(row=3, column=9).border = thin_border
        ws.cell(row=3, column=9).alignment = Alignment(horizontal='center')
        i3=ws['I3']
        i3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        ws['J3'] = 'LUGAR'
        ws.cell(row=3, column=10).border = thin_border
        ws.cell(row=3, column=10).alignment = Alignment(horizontal='center')
        j3=ws['J3']
        j3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        cont = 4
        indice = 1
        for auditoria in auditorias_list:



            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = auditoria['numero_auditoria']
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = auditoria['nombre_norma']
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=4).value = auditoria['nombre_area']
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=5).value = auditoria['nombre_proceso']
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')


            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=6).value = auditoria['fec_inicio']
            ws.cell(row=cont,column=6).border = thin_border
            ws.cell(row=cont,column=6).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=7).value = auditoria['hora_inicio']
            ws.cell(row=cont,column=7).border = thin_border
            ws.cell(row=cont,column=7).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=8).value = auditoria['fec_fin']
            ws.cell(row=cont,column=8).border = thin_border
            ws.cell(row=cont,column=8).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=9).value = auditoria['nombre_auditor']
            ws.cell(row=cont,column=9).border = thin_border
            ws.cell(row=cont,column=9).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=10).value = auditoria['lugar']
            ws.cell(row=cont,column=10).border = thin_border
            ws.cell(row=cont,column=10).alignment = Alignment(horizontal='center')


            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Listado_Auditorias.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["A"].width = 5.0
        ws.column_dimensions["B"].width = 20.0
        ws.column_dimensions["C"].width = 20.0
        ws.column_dimensions["D"].width = 30.0
        ws.column_dimensions["E"].width = 30.0
        ws.column_dimensions["F"].width = 30.0
        ws.column_dimensions["G"].width = 20.0
        ws.column_dimensions["H"].width = 20.0
        ws.column_dimensions["I"].width = 20.0
        ws.column_dimensions["J"].width = 40.0
        ws.column_dimensions["K"].width = 40.0

        wb.save(response)
        #retorna el archivo excel
        return response


def pdfAudits(request):

        
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #Consultar
        cursor = connection.cursor()
        """
        query = "SELECT auditoria_auditorias.id AS id_auditoria, auditoria_auditorias.lugar,  auditoria_auditorias.fec_inicio, auditoria_auditorias.fec_fin, "
        query = query + " auditoria_auditorias.objetivo, auditoria_auditorias.id_area_id, auditoria_auditorias.id_norma_id, "
        query = query + " auditoria_auditorias.hora_inicio, auditoria_auditorias.numero_auditoria, auditoria_auditorias.id_auditor_id, "
        query = query + " auditoria_auditorias.id_proceso_id, auditoria_auditorias.numero_auditoria, "
        query = query + " normas_normas.nombre AS nombre_norma, "
        query = query + " areas_areas.area AS nombre_area, "
        query = query + " areas_procesos.proceso AS nombre_proceso, "
        query = query + " areas_subproceso.nombre AS subproceso, "
        query = query + " personal_personal.nombre AS nombre_auditor "
        query = query + " FROM auditoria_auditorias "
        query = query + " LEFT JOIN normas_normas on auditoria_auditorias.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN areas_areas on auditoria_auditorias.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos on auditoria_auditorias.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN areas_subproceso ON auditoria_auditorias.id_sub_proceso_id = areas_subproceso.id_proceso_id "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditorias.id_auditor_id = personal_personal.id"
        query = query + " WHERE areas_subproceso.nombre IS NOT NULL"
        """
        query = "SELECT auditoria_auditorias.id AS id_auditoria, auditoria_auditorias.lugar,  auditoria_auditorias.fec_inicio, auditoria_auditorias.fec_fin, "
        query = query + " auditoria_auditorias.objetivo, auditoria_auditorias.id_area_id, auditoria_auditorias.id_norma_id, "
        query = query + " auditoria_auditorias.hora_inicio, auditoria_auditorias.numero_auditoria, auditoria_auditorias.id_auditor_id, "
        query = query + " auditoria_auditorias.id_proceso_id, auditoria_auditorias.numero_auditoria, "
        query = query + " normas_normas.nombre AS nombre_norma, "
        query = query + " areas_areas.area AS nombre_area, "
        query = query + " areas_procesos.proceso AS nombre_proceso, "
        query = query + " personal_personal.nombre AS nombre_auditor "
        query = query + " FROM auditoria_auditorias "
        query = query + " LEFT JOIN normas_normas on auditoria_auditorias.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN areas_areas on auditoria_auditorias.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos on auditoria_auditorias.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN personal_personal ON auditoria_auditorias.id_auditor_id = personal_personal.id"


        cursor.execute(query)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        auditorias_list = list(rows) 

        context = {
            'auditorias_list': auditorias_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfAuditorias.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Reporte_Auditorias_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response
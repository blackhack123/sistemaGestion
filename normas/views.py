# -*- coding: utf-8 -*-
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
import json
#para upload
from django.core.files.storage import FileSystemStorage
import os
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#IMPORTAMOS BORDES 
from openpyxl.styles import Border, Side, Alignment, Color, Fill, colors, Font, PatternFill
from openpyxl.cell import Cell
#importamos render_to_pdf de vista principal
from principal.views import render_to_pdf
#IMPORTAMOS LA CONEXION PARA QUERY PERSONALIZADO
from django.db import connection
#importamos el convert diccionario
from principal.views import dictfetchall
#librerias pdfminer
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter, HTMLConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from cStringIO import StringIO
import re


#importamos models-views-forms
from normas.models import Normas,Clausulas, ProcesoClausula    
from normas.forms import FormNorma
from areas.views import Areas
from areas.models import Procesos
#importamos los cargos
from personal.models import Cargo, Personal
# Create your views here.

#INICIO DE PROCESOS Y SU VINCULACION
def normas(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        
        context={
            'formNorma':FormNorma(),
            'cargos':Cargo.objects.all().filter(estado=1)
        }
        return render(request, 'admin/normas.html', context)


def gridNormas(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

            # todos los valores
            #normas = Normas.objects.all().values('id', 'nombre', 'estado','docfile')
            cursor = connection.cursor()

            query = " SELECT normas_normas.id, normas_normas.nombre, normas_normas.estado, normas_normas.docfile, "
            query = query + " personal_personal.nombre AS auditor_lider, personal_personal.id_cargo_id, personal_personal.id AS id_personal "
            query = query + " FROM normas_normas "
            query = query + " LEFT JOIN personal_personal ON normas_normas.auditor_lider_id = personal_personal.id "
            cursor.execute(query)

            rows = dictfetchall(cursor)

            # CONVIERTE  QuerySet  a list object
            normas_list = list(rows)  

            if normas_list:
                #RETORNA  
                return JsonResponse(normas_list, safe=False)
            else:
                #RETORNA MENSAJE DE NO EXISTEN
                json = {
                'resultado':'no_ok',
                'mensaje': 'no existen Normas',
                }
                return JsonResponse(json, safe=False)


def selectNorma(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_norma = request.POST.get('id_norma')

            norma = Normas.objects.get(pk=id_norma)
            cursor = connection.cursor()

            query = " SELECT normas_normas.id,  normas_normas.nombre, normas_normas.estado, normas_normas.docfile, personal_personal.nombre AS auditor_lider, "
            query = query + " personal_personal.id_cargo_id,  personal_personal.id AS id_personal "
            query = query + " FROM normas_normas "
            query = query + " LEFT JOIN personal_personal ON normas_normas.auditor_lider_id = personal_personal.id "
            query = query + " WHERE normas_normas.id =%s "
            params=[id_norma]

            cursor.execute(query, params)
            rows = dictfetchall(cursor)
            norma_list = list(rows)
            """
            urlDocumento= str(norma.docfile)
            norma_list ={
                'id':norma.id,
                'nombre':norma.nombre ,
                'docfile':urlDocumento,
                'estado':norma.estado,
            }
            """
            data = {
            'resultado':'ok_select',
            'norma_list':norma_list,
            }
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})


def insertNorma(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        formNorma = FormNorma(request.POST, request.FILES)
        if request.method == 'POST' and request.FILES:

            myfile = request.FILES['docNorma']
            fs = FileSystemStorage(location='media/normas')
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = 'media/normas/'+filename
            
            norma = Normas.objects.filter(nombre=request.POST.get("norma")).count()
            if norma == 0:
                normas = Normas(
                    docfile=uploaded_file_url,
                    nombre=request.POST.get("norma"),
                    estado=request.POST.get("estado"),
                    auditor_lider_id=request.POST.get('personal')
                )
                normas.save()

                return redirect('normas')
            else:
                from django.contrib import messages
                messages.add_message(request, messages.WARNING, 'La norma ya existe.. no puede ser creada !!')
            
                return redirect('normas')
        else:
            norma = Normas.objects.filter(nombre=request.POST.get("norma")).count()
            if norma == 0:
                normas = Normas(
                    nombre=request.POST.get("norma"),
                    estado=request.POST.get("estado"),
                    auditor_lider_id=request.POST.get('personal')
                )
                normas.save()
                return redirect('normas')
            else:

                from django.contrib import messages
                messages.add_message(request, messages.WARNING, 'La norma ya existe.. no puede ser creada !!')
            
                return redirect('normas')



def updateNorma(request):


    if 'nombreUsuario' not in request.session:

            return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        form = FormNorma(request.POST, request.FILES)
        if request.method == 'POST' and request.FILES:

            myfile = request.FILES['docNorma']
            fs = FileSystemStorage(location='media/normas')
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = 'media/normas/'+filename
            
            Normas.objects.filter(pk=request.POST.get('idNorma')).update(
                docfile=uploaded_file_url,
                nombre=request.POST.get("norma"),
                estado=request.POST.get("estado"),
                auditor_lider_id=request.POST.get('personal')
            )
            return redirect('normas')
        else:
            
            Normas.objects.filter(pk=request.POST.get('idNorma')).update(
                nombre=request.POST.get("norma"),
                estado=request.POST.get("estado"),
                auditor_lider_id=request.POST.get('personal')
            )
            return redirect('normas')


def excelNormas(request):
    
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        #Consultar
        normas=Normas.objects.all().values('id', 'nombre', 'estado')
        # CONVIERTE  QuerySet  a list object
        normas_list = list(normas)  

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Normas"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:C2')
        ws['B2'] = 'Reporte General de Normas'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        ws['C3'] = 'Norma'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        cont = 4
        indice = 1
        for norma in normas_list:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')
            

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = norma['nombre']
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

                
            if norma['estado'] == "0":

                #obtengo celdas de estado inactivo
                normaCell=ws.cell(row=cont,column=3)

                #color rojo a inactivos
                normaCell.font = Font(color=colors.RED)

            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Reporte_General_Norma.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0

        wb.save(response)
        #retorna el archivo excel
        return response

def pdfNormas(request):
    
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
      
        #Consultar
        normas=Normas.objects.all().values('id', 'nombre', 'estado')
        # CONVIERTE  QuerySet  a list object
        normas_list = list(normas)  

        context = {
            'normas': normas_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfNormas.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Reporte_Normas_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response



def vincularNormas(request):

    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        context={
            'normas':Normas.objects.all().values('id','nombre').filter(estado=1),
            'areas':Areas.objects.all().values('id','area').filter(estado=1),
        }
        return render(request, 'admin/vincularNormas.html', context)



def gridVincularProcesos(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        # or simply .values() to get all fields
        id_area=request.GET["id_area"]
        #procesos = Procesos.objects.all().values('id','proceso').filter(estado=1, id_area_id=id_area)
        
        cursor = connection.cursor()
        query = "SELECT DISTINCT areas_area_proceso.proceso_id AS id,"
        query = query + " areas_procesos.proceso "
        query = query + " FROM areas_area_proceso "
        query = query + " LEFT JOIN areas_procesos ON areas_area_proceso.proceso_id = areas_procesos.id "
        query = query + " WHERE areas_area_proceso.area_id = %s "
        params=[id_area]
        #ejecutar consulta
        cursor.execute(query, params)
        rows=dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        procesos_list = list(rows)  

        if procesos_list:
            #RETORNA  
            return JsonResponse(procesos_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Procesos',
            }
            return JsonResponse(json, safe=False)




def gridVinClausulas(request):


    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        # or simply .values() to get all fields
        id_norma=request.GET["id_norma"]
        clausulas = Clausulas.objects.all().values('id','clausula').filter(id_norma_id=id_norma)
        
        # CONVIERTE  QuerySet  a list object
        clausulas_list = list(clausulas)  

        if clausulas_list:
            #RETORNA  
            return JsonResponse(clausulas_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Clausulas',
            }
            return JsonResponse(json, safe=False)




def insertVinculoNormas(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            #DATA RETORNADA DEL AJAX
            procesos = json.loads(request.POST.get('data'))
       
                #norma=Normas.objects.get(pk=proceso['id_norma'])
            try:

                for proceso in procesos['gridData']:
                    procesoClausula = ProcesoClausula.objects.get(id_proceso_id=proceso['id_proceso'], id_clausula_id=proceso['id_clausula'])
                    if procesoClausula:
                        procesoExistente=Procesos.objects.get(pk=proceso['id_proceso'])
                        data = {
                        'resultado':'no_ok',
                        'proceso':procesoExistente.proceso,
                        }
                        return HttpResponse(json.dumps(data), content_type="application/json")

            except ProcesoClausula.DoesNotExist :

                for proceso in procesos['gridData']:

                    procesoClausula = ProcesoClausula(
                        id_area_id=proceso['id_area'],
                        id_clausula_id=proceso['id_clausula'],
                        id_norma_id=proceso['id_norma'],
                        id_proceso_id=proceso['id_proceso'],
                    )
                    procesoClausula.save()
                    
                #retornamos la respuesta
                data = {
                'resultado':'ok_insert',
                'mensaje': 'Datos Guardados Correctamente !!',
                }
                return HttpResponse(json.dumps(data), content_type="application/json")
        else:

            return render(request, 'seguridad/login.html', {})



def gridVinculadoNormas(request):
    
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        cursor = connection.cursor()
        query = " SELECT  areas_areas.area, areas_procesos.proceso,  normas_normas.nombre AS nombre_norma, normas_clausulas.clausula, "
        query = query + " normas_procesoclausula.id, normas_procesoclausula.id_area_id, normas_procesoclausula.id_clausula_id, "
        query = query + "  normas_procesoclausula.id_norma_id, normas_procesoclausula.id_proceso_id "
        query = query + " FROM normas_procesoclausula"
        query = query + " LEFT JOIN areas_areas ON normas_procesoclausula.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON normas_procesoclausula.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN normas_normas ON normas_procesoclausula.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN normas_clausulas ON normas_procesoclausula.id_clausula_id = normas_clausulas.id "
        query = query + " ORDER BY nombre_norma, clausula "
        cursor.execute(query)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        normas_list = list(rows)  

        if normas_list:
            #RETORNA  
            return JsonResponse(normas_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Normas Vinculadas',
            }
            return JsonResponse(json, safe=False)



def desvincularNormaProceso(request):


    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            id_proceso = request.POST.get("id_proceso")
            #desvincular clausula del proceso
            ProcesoClausula.objects.filter(pk=id_proceso).delete()
            data = {
                'resultado':'ok_update',
                'mensaje': 'Clausula Desvinculada Exitosamente !!',
            }
                
            return HttpResponse(json.dumps(data), content_type="application/json")



def desvincularNormasProcesos(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            id_norma = request.POST.get("id_norma")
            id_proceso = request.POST.get("id_proceso")
            #desvincular todo los procesos 
            ProcesoClausula.objects.filter(id_norma_id=id_norma, id_proceso_id=id_proceso).delete()

            data = {
                'resultado':'ok_update',
                'mensaje': 'Procesos Desvinculados Correctamente !!',
            }
                
            return HttpResponse(json.dumps(data), content_type="application/json")



def excelNormVinculadas(request):
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        cursor = connection.cursor()
        query = " SELECT  areas_areas.area, areas_procesos.proceso,  normas_normas.nombre AS nombre_norma, normas_clausulas.clausula "
        query = query + " FROM normas_procesoclausula"
        query = query + " LEFT JOIN areas_areas ON normas_procesoclausula.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON normas_procesoclausula.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN normas_normas ON normas_procesoclausula.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN normas_clausulas ON normas_procesoclausula.id_clausula_id = normas_clausulas.id "
        query = query + " ORDER BY normas_procesoclausula.id_proceso_id"

        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        normas_list = list(rows)  

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Normas Vinculadas"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:F2')
        ws['B2'] = 'Reporte General de Normas Vinculadas'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        ws['C3'] = 'Norma'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Proceso'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Procedimiento'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['F3'] = 'Cláusula'
        ws.cell(row=3, column=6).border = thin_border
        ws.cell(row=3, column=6).alignment = Alignment(horizontal='center')
        f3=ws['F3']
        f3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for norma in normas_list:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')
            

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = norma['nombre_norma']
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=4).value = norma['area']
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=5).value = norma['proceso']
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=6).value = norma['clausula']
            ws.cell(row=cont,column=6).border = thin_border
            ws.cell(row=cont,column=6).alignment = Alignment(horizontal='center')

            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Reporte_Norma_Vinculadas.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 40.0
        ws.column_dimensions["D"].width = 30.0
        ws.column_dimensions["E"].width = 30.0
        ws.column_dimensions["F"].width = 60.0

        wb.save(response)
        #retorna el archivo excel
        return response


def pdfNormVinculadas(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        cursor = connection.cursor()
        query = " SELECT  areas_areas.area, areas_procesos.proceso,  normas_normas.nombre AS nombre_norma, normas_clausulas.clausula "
        query = query + " FROM normas_procesoclausula"
        query = query + " LEFT JOIN areas_areas ON normas_procesoclausula.id_area_id = areas_areas.id "
        query = query + " LEFT JOIN areas_procesos ON normas_procesoclausula.id_proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN normas_normas ON normas_procesoclausula.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN normas_clausulas ON normas_procesoclausula.id_clausula_id = normas_clausulas.id "
        query = query + " ORDER BY normas_procesoclausula.id_proceso_id"

        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        normas_list = list(rows)  

        context = {
            'normas': normas_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfNormasVinculadas.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Reporte_Normas_Vinculadas_%s.pdf" %("0001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response


def clausulas(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        cursor = connection.cursor()
        query = "SELECT DISTINCT normas_normas.id AS id_norma, normas_normas.nombre AS norma, normas_clausulas.clausula,  normas_clausulas.detalle, normas_clausulas.id, "
        query = query + " CASE WHEN normas_procesoclausula.id_clausula_id IS NULL THEN 0 ELSE 1 END AS flag "
        query = query + " FROM normas_clausulas "
        query = query + " LEFT JOIN normas_normas ON normas_clausulas.id_norma_id = normas_normas.id "
        query = query + " LEFT JOIN normas_procesoclausula ON normas_clausulas.id = normas_procesoclausula.id_clausula_id "

        cursor.execute(query)
        rows = dictfetchall(cursor)
        clausulas_list = list(rows)
        context={
           'normas':Normas.objects.values('id', 'nombre').all().filter(estado=1),
           'clausulas':clausulas_list,
        }
        return render(request, 'admin/clausulas.html', context)




def selectClausula(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:

        if request.method == 'POST' and request.is_ajax():

            id_clausula=request.POST.get('id_clausula')

            clausula=Clausulas.objects.get(pk=id_clausula)
            
            clausula_list ={
                'id':clausula.id,
                'clausula':clausula.clausula ,
                'descripcion':clausula.detalle,
                'id_norma':clausula.id_norma_id,
            }
    
            data = {
            'resultado':'ok_select',
            'clausula_list':clausula_list,
            }
            return HttpResponse(json.dumps(data), content_type="application/json")



def convert_scanned_to_pdf(path, numero_pagina):
    
    """
    funcion que permite convertir la pagina deseada
    a imagen y esta a texto, pasando por filtro 
    de mejoramiento.
    Se debera pasar dos parametros:
    path = url de la ubicacion del archivo (string)
    numero_pagina = pagina a convertir a texto (integer)
    """
    
    #convertir numero_pagina a int
    numero = int(numero_pagina)

    #asignar el path a una variable
    path_absoluta = path

    #armar la path con numero de pagina
    path_absoluta = path_absoluta + "["+str(numero-1)+"]"

    #verificar si existe el archivo temporal
    from pathlib import Path
    temp_file = Path("media/normas/temp.jpg")
    sample_file = Path("media/normas/sample.jpeg")
    if temp_file.is_file():
        #remover archivo temporal
        import os
        os.remove("media/normas/temp.jpg")

    if sample_file.is_file():
        #remover archivo temporal
        import os
        os.remove("media/normas/sample.jpeg")


    #convertir pdf a imagen
    from wand.image import Image
    #estableciendo resolucion a imagen
    with Image(filename=path_absoluta, resolution=400) as img:
        #estableciendo ancho y alto
        img.resize(1850,1850)
        img.save(filename="media/normas/temp.jpg")
    

    import pytesseract
    from PIL import ImageEnhance, ImageFilter
    from PIL import Image as Img
    im = Img.open("media/normas/temp.jpg")
    im = im.filter(ImageFilter.MedianFilter())

    enhancer = ImageEnhance.Contrast(im)
    #aplicando filtro para mejorar la convercion de imagen->txt
    im = enhancer.enhance(15)
    im = im.convert('1')
    im.save('media/normas/sample.jpeg')
    
    
    text = pytesseract.image_to_string(Img.open('media/normas/sample.jpeg'), lang='spa')
   
    #grabando el texto en archivo temporal
    text_file = open("media/normas/temporalPdf.txt", "w")
    text = re.sub("\s\s+", " ", text)

    from django.utils.encoding import smart_str, smart_unicode
    text_file.write("%s" % smart_str(text))
    text_file.close()

    

def convert_pdf_to_txt(path, numero_pagina):

    """
    Funcion que permite convertir una pagina de pdf 
    a formato html(p, tables, divs, etc..)
    con el fin de mantener estetico el documento
    Se pasaran dos parametros
    path = url de la ubicacion del archivo(string)
    numero_pagina = pagina a convertir a texto(integer)
    """

    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    #device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    device = HTMLConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = file(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos=set()

    #for pageNumber, page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
    for pageNumber, page in enumerate(PDFPage.get_pages(fp, pagenos)):
        if pageNumber == int(numero_pagina)-1:
            interpreter.process_page(page)

    text = retstr.getvalue()
 
    
    fp.close()
    device.close()
    retstr.close()

    text_file = open("media/normas/temporalPdf.txt", "w")
    text = re.sub("\s\s+", " ", text)
    text_file.write("%s" % text)
    text_file.close()
    


def extraerClausula(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            id_norma=request.POST.get('id_norma')
            tipo_pdf = request.POST.get('escaneado_pdf')
            numero_pagina = request.POST.get('numero_pagina')
            #se obtiene la norma
            norma = Normas.objects.get(pk=id_norma)
            #se obtiene la url del pdf
            path= str(norma.docfile)
            #convert pdf a txt
            if tipo_pdf == '1':
                #si PDF escaneado
                convert_scanned_to_pdf(path, numero_pagina)

            else:
                #si es PDF normal
                convert_pdf_to_txt(path, numero_pagina)

            #convert_pdf_to_txt(path)
            data = {
                'resultado':'ok_select',
            }
            
            return HttpResponse(json.dumps(data), content_type="application/json")



def insertClausula(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        if request.method == 'POST' and request.is_ajax():

            id_norma=request.POST.get('id_norma')
            clausulaNombre=request.POST.get('clausula')
            descripcion=request.POST.get('descripcion')

            clausula_consultar = Clausulas.objects.filter(clausula=clausulaNombre, id_norma_id=id_norma).count()
            if clausula_consultar == 0:
                clausula=Clausulas(
                    id_norma_id=id_norma,
                    clausula=clausulaNombre,
                    detalle=descripcion.encode('unicode_escape')
                )
                clausula.save()

                data = {
                'resultado':'ok_insert',
                'mensaje': 'Datos Guardados Correctamente !!',
                }
            else:
                data = {
                    'resultado':'no_ok',
                    'mensaje': 'La Clausula ya existe no puede ser creada !!',
                }
            return HttpResponse(json.dumps(data), content_type="application/json")
        else:

            return render(request, 'seguridad/login.html', {})



def updateClausula(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST' and request.is_ajax():
            id_clausula=request.POST.get('id_clausula')
            id_norma=request.POST.get('id_norma')
            clausulaNombre=request.POST.get('clausula')
            descripcion=request.POST.get('descripcion')

            Clausulas.objects.filter(pk=id_clausula).update(
                clausula=clausulaNombre,
                id_norma_id=id_norma,
                detalle=descripcion,
            )
            data = {
                'resultado':'ok_update',
                'mensaje': 'Datos Actualizados Correctamente !!',
            }
            return HttpResponse(json.dumps(data), content_type="application/json")


def consultarProcesoNorma(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        if request.method == 'POST' and request.is_ajax():

            id_clausula = request.POST.get('id_clausula')

            #conexion
            cursor = connection.cursor()

            #query
            query = " SELECT areas_procesos.proceso "
            query = query + " FROM normas_procesoclausula "
            query = query + " LEFT JOIN areas_procesos ON normas_procesoclausula.id_proceso_id = areas_procesos.id "
            query = query + " WHERE normas_procesoclausula.id_clausula_id =%s "

            #parametros
            params =[id_clausula]
            #ejecutar consulta
            cursor.execute(query, params)

            #returnar diccionar de consulta
            rows = dictfetchall(cursor)

            # CONVIERTE  QuerySet  a list object
            procesos_list = list(rows)

            #recorriendo los directorios
            alldirectorys = []
            import os
            for directorios in procesos_list:
                for root, dirs, files in os.walk('media/gestionDocumental'):
                    for directory in dirs:
                        if directory == directorios['proceso']:


                            element = {}
                            element['directorio'] = directory
                            element['ruta_absoluta'] = root + '/' + directory
                            element['archivos'] = os.listdir(root+'/'+directorios['proceso'])
                            alldirectorys.append(element)

            if procesos_list:
                data = {
                    'resultado':'ok_select',
                    'procesos_list':procesos_list,
                    'archivos':alldirectorys
                }
                #RETORNA  
                return JsonResponse(data, safe=False)
            else:
                #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
                json = {
                'resultado':'no_ok',
                'mensaje': 'no existen Procesos Vinculados'
                }
                return JsonResponse(json, safe=False)



def ExcelClausulas(request):
    if 'nombreUsuario' not in request.session:
        return render(request, 'seguridad/login.html', {})
    else:
        cursor = connection.cursor()
        query = " SELECT normas_normas.nombre AS norma, normas_clausulas.clausula, normas_clausulas.detalle "
        query = query + " FROM normas_clausulas "
        query = query + " LEFT JOIN normas_normas ON normas_clausulas.id_norma_id = normas_normas.id "
        cursor.execute(query)

        clausulas_list = dictfetchall(cursor)

        # ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Creamos el libro de trabajo
        wb = Workbook()

        # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Clausulas"
        # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Reporte General de Clausulas'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2 = ws['B2']
        b2.font = Font(bold=True, color=colors.DARKBLUE, size=12)
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3 = ws['B3']
        b3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'Norma'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3 = ws['C3']
        c3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'Clausula'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3 = ws['D3']
        d3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'Detalle'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3 = ws['E3']
        e3.font = Font(bold=True, color=colors.DARKBLUE, size=12)

        cont = 4
        indice = 1
        for clausula in clausulas_list:

            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=2).value = indice
            ws.cell(row=cont, column=2).border = thin_border
            ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')

            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=3).value = clausula['norma']
            ws.cell(row=cont, column=3).border = thin_border
            ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=4).value = clausula['clausula']
            ws.cell(row=cont, column=4).border = thin_border
            ws.cell(row=cont, column=4).alignment = Alignment(horizontal='center')

            # agregamos la data, borde, alineacion
            ws.cell(row=cont, column=5).value = clausula['detalle']
            ws.cell(row=cont, column=5).border = thin_border
            ws.cell(row=cont, column=5).alignment = Alignment(horizontal='center')


            cont = cont + 1
            indice = indice + 1

        # Establecemos el nombre del archivo
        nombre_archivo = "Reporte_General_clausulas.xlsx"

        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        # ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 40.0
        ws.column_dimensions["E"].width = 40.0
        wb.save(response)
        # retorna el archivo excel
        return response
# -*- coding: utf-8 -*-
import os
import shutil

from django.shortcuts import render, redirect
from areas.models import Areas, Procesos, Area_Proceso
from django.http import HttpResponse, JsonResponse, HttpResponseNotFound
import json
from django.contrib import messages
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#IMPORTAMOS BORDES 
from openpyxl.styles import Border, Side, Alignment, Color, Fill, colors, Font, PatternFill
from openpyxl.cell import Cell
#importamos render_to_pdf de vista principal
from principal.views import render_to_pdf
#IMPORTAMOS LA CONEXION PARA QUERY PERSONALIZADO
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.db import connection

from procesos.forms import ProcesoForm
from documentos.forms import DocumentacionForm
from personal.models import Personal, Cargo
from procesos.models import Colaboradores
from documentos.models import Documento, Revision
from seguridad.models import Usuario


#INICIO DE PROCESOS Y SU VINCULACION
def procesos(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        #LISTADO DE SOLO DIRECTORIOS
        alldirectorys = []
        rootDir = 'media/gestionDocumental/'
        for dirName, subdirList, fileList in os.walk(rootDir):
            #validamos que vaya la root
            if dirName != 'media/gestionDocumental/':
                
                element={}
                #ruta_absoluta
                element['ruta_absoluta']=dirName
                #se hace split para mostrar solo el nombre del folder no completo
                element['nombre']= dirName.split('media/gestionDocumental/')
                alldirectorys.append(element)

        context={
            'areas':Areas.objects.all(),
            #'procesos':Procesos.objects.values('id', 'proceso', 'estado').filter(estado=1),
            'formProceso':ProcesoForm(),
            'directorios':alldirectorys
        }
        return render(request, 'admin/procesos.html', context)


def dictfetchall(cursor):

    "Retorna data como diccionario y con keys"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]


def gridProcesos(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        cursor = connection.cursor()
        query = " SELECT DISTINCT a.id AS id_proceso, a.proceso,  a.tipo_proceso, a.estado AS estado_proceso, "
        query = query + " CASE WHEN b.proceso_id IS NULL THEN 0 ELSE 1 END AS flag "
        query = query + " FROM areas_procesos a "
        query = query + " LEFT JOIN areas_area_proceso b ON a.id = b.proceso_id  ORDER BY proceso "

        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        procesos_list = list(rows)  

        if procesos_list:
            #RETORNA  
            return JsonResponse(procesos_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Procesos',
            }
            return JsonResponse(json, safe=False)



def consultarAreas(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_proceso = request.POST.get('id_proceso')
            
            #conexion
            cursor = connection.cursor()
            #consulta select areas por id_proceso
            query = " SELECT DISTINCT areas_areas.area "
            query = query + " FROM areas_area_proceso "
            query = query + " LEFT JOIN areas_areas ON areas_area_proceso.area_id = areas_areas.id "
            query = query + " WHERE areas_area_proceso.proceso_id = %s "
            query = query + " GROUP BY areas_area_proceso.area_id "

            #parametro
            params = [id_proceso]
            cursor.execute(query, params)
            
            rows = dictfetchall(cursor)
            areas_list = list(rows)  

            if areas_list:
                #RETORNA  
                return JsonResponse(areas_list, safe=False)
            else:
                #RETORNA MENSAJE DE NO EXISTEN
                json = {
                    'resultado':'no_ok',
                    'mensaje': 'no existen Areas Vinculadas !!',
                }
                return JsonResponse(json, safe=False)



def selectProceso(request):


    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_proceso = request.POST.get('id_proceso')

            proceso = Procesos.objects.get(pk=id_proceso)
            stringDoc=str(proceso.docfile)
            proceso_list ={
                'id':proceso.id,
                'nombre':proceso.proceso ,
                'docfile':stringDoc,
                'tipo':proceso.tipo_proceso,
                'estado':proceso.estado,
                'url_carpeta':proceso.url_carpeta
            }
    
            data = {
            'resultado':'ok_select',
            'proceso_list':proceso_list,
            }
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:

            return render(request, 'seguridad/login.html', {})  




def insertProceso(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        #validar si recibe los datos por ajax
        form = ProcesoForm(request.POST, request.FILES)
        if request.method == 'POST' and request.FILES:
            

            myfile = request.FILES['docProceso']
            destino_archivo = request.POST['destino']+'/'+request.POST['directorio']
            fs = FileSystemStorage(location=destino_archivo)
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = 'media/procesos/'+filename

            proceso = Procesos(
                docfile=uploaded_file_url,
                proceso=form['proceso'].value(),
                tipo_proceso=form['tipo_proceso'].value(),
                disponible=1,
                estado=form['estado'].value(),
                url_carpeta=str(request.POST['destino'])+'/'+str(request.POST['directorio'])
            )
            proceso.save()

            messages.add_message(request, messages.SUCCESS, 'Registro Creado Existosamente !!')
            return redirect('/procesos')
            
        else:
            proceso_existente = Procesos.objects.filter(proceso=form['proceso'].value()).count()
            if proceso_existente == 0:
                proceso = Procesos(
                    proceso=form['proceso'].value(),
                    tipo_proceso=form['tipo_proceso'].value(),
                    estado=form['estado'].value(),
                    url_carpeta=str(request.POST['destino'].encode('utf-8'))+'/'+request.POST['directorio'].encode('utf-8')
                )
                proceso.save()
                #creando la carpeta de destino
                try:
                    os.makedirs(str(request.POST['destino'].encode('utf-8'))+"/"+str(request.POST['directorio'].encode('utf-8')), mode=0777)
                except OSError:
                    pass
                messages.add_message(request, messages.SUCCESS, 'Registro Creado Existosamente !!')
                return redirect('/procesos')
            else:
                messages.add_message(request, messages.WARNING, 'El registro ya existe.. no puede ser creado !!')
                return redirect('procesos')


def updateProceso(request):

    if 'nombreUsuario' not in request.session:

            return render(request, 'seguridad/login.html', {})

    else:
        form = ProcesoForm(request.POST, request.FILES)
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.FILES:


            myfile = request.FILES['docProceso']
            fs = FileSystemStorage(location='media/procesos')
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = 'media/procesos/'+filename

            Procesos.objects.filter(pk=request.POST.get('idProceso')).update(
                proceso= request.POST.get('proceso'), 
                estado=request.POST.get('estado'),
                tipo_proceso=form['tipo_proceso'].value(),
                docfile=uploaded_file_url,
            )
            return redirect('procesos')
        else:
            
            Procesos.objects.filter(pk=request.POST.get('idProceso')).update(
                proceso= request.POST.get('proceso'), 
                tipo_proceso=form['tipo_proceso'].value(),
                estado=request.POST.get('estado'),
            )
            
            #renombrar carpeta

            carpeta_anterior = str(request.POST.get('carpeta_anterior'))
            nuevo_nombre = str(request.POST.get('proceso'))
            if carpeta_anterior != nuevo_nombre:
                try:
                    path = 'media/gestionDocumental/'
                    path_completo = str(request.POST.get('destino'))
                    for root, dirs, files in os.walk(path):
                        for directory in dirs:
                            if directory == carpeta_anterior:
                                os.rename(str(path_completo+'/'+carpeta_anterior), str(path_completo+'/'+nuevo_nombre))
                                messages.add_message(request, messages.SUCCESS, 'Registro Actualizado Existosamente !!')
                                return redirect('procesos')
                except OSError:
                    pass
                    return redirect('procesos')
            else:
                messages.add_message(request, messages.SUCCESS, 'Registro Actualizado Existosamente !!')
                return redirect('procesos')

    
                
     

def excelProcesos(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        #Consultar
        cursor = connection.cursor()
        query= "SELECT areas_areas.area AS proceso, personal_personal.nombre AS encargado_proceso, areas_procesos.proceso AS procedimiento, areas_procesos.estado AS estado_procedimiento "
        query = query + " FROM areas_procesos"
        query = query + " LEFT JOIN areas_area_proceso ON areas_area_proceso.proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN areas_areas ON areas_area_proceso.area_id = areas_areas.id "
        query = query + " LEFT JOIN personal_personal ON areas_areas.id_personal_id = personal_personal.id " 
        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        procesos_list = list(rows)  
        
        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Procedimientos"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Reporte General de Procedimientos'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True, color=colors.DARKBLUE, size=12)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['C3'] = 'PROCESO'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['D3'] = 'ENCARGADO DEL PROCESO'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True, color=colors.DARKBLUE, size=12)

        ws['E3'] = 'PROCEDIMIENTO'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True, color=colors.DARKBLUE, size=12)


        cont = 4
        indice = 1
        for proceso in procesos_list:

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')
            
            #agregamos la data, borde, alineacion
            if proceso['proceso']:
                ws.cell(row=cont,column=3).value = proceso['proceso']
            else:
                ws.cell(row=cont,column=3).value = "--"
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            if proceso['encargado_proceso']:
                ws.cell(row=cont,column=4).value = proceso['encargado_proceso']
            else:
                ws.cell(row=cont,column=4).value = "--"
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=5).value = proceso['procedimiento']
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')

                
            if proceso['estado_procedimiento'] == "0":

                #obtengo celdas de estado inactivo
                areaCell=ws.cell(row=cont,column=3)
                jefeCell=ws.cell(row=cont,column=4)
                procesoCell=ws.cell(row=cont,column=5)

                #color rojo a inactivos
                areaCell.font = Font(color=colors.RED)
                jefeCell.font = Font(color=colors.RED)
                procesoCell.font = Font(color=colors.RED)

            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Reporte_General_Procesos.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 30.0
        ws.column_dimensions["E"].width = 30.0

        wb.save(response)
        #retorna el archivo excel
        return response


def pdfProcesos(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
      
        #Consultar
        cursor = connection.cursor()
        query= "SELECT areas_areas.area AS proceso, personal_personal.nombre AS encargado_proceso, areas_procesos.proceso AS procedimiento, areas_procesos.estado AS estado_procedimiento "
        query = query + " FROM areas_procesos"
        query = query + " LEFT JOIN areas_area_proceso ON areas_area_proceso.proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN areas_areas ON areas_area_proceso.area_id = areas_areas.id "
        query = query + " LEFT JOIN personal_personal ON areas_areas.id_personal_id = personal_personal.id " 
        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        procesos_list = list(rows)   

        context = {
            'procesos': procesos_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfProcesos.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Reporte_Procesos_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response
      


def vincularProcesos(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:
        #conexion
        cursor = connection.cursor()

        #query directores
        query = " SELECT personal_personal.id, personal_personal.nombre "
        query = query + " FROM seguridad_usuario "
        query = query + " LEFT JOIN personal_personal ON seguridad_usuario.id_personal_id = personal_personal.id "
        query = query + " WHERE seguridad_usuario.director = 1 AND seguridad_usuario.estado = 1 "
        cursor.execute(query)
        rows_director = dictfetchall(cursor)

        #query colaboradores
        query = " SELECT personal_personal.id, personal_personal.nombre "
        query = query + " FROM seguridad_usuario "
        query = query + " LEFT JOIN personal_personal ON seguridad_usuario.id_personal_id = personal_personal.id "
        query = query + " WHERE seguridad_usuario.colaborador = 1 AND seguridad_usuario.estado = 1 "
        cursor.execute(query)
        rows_colaborador = dictfetchall(cursor)

        # query lider_norma
        query = " SELECT personal_personal.id, personal_personal.nombre "
        query = query + " FROM seguridad_usuario "
        query = query + " LEFT JOIN personal_personal ON seguridad_usuario.id_personal_id = personal_personal.id "
        query = query + " WHERE seguridad_usuario.lider_norma = 1 AND seguridad_usuario.estado = 1 "
        cursor.execute(query)
        rows_lider_norma = dictfetchall(cursor)

        context={
            'areas':Areas.objects.all().filter(estado=1),
            'cargos':Cargo.objects.values('id', 'nombre').all().filter(estado=1),
            'procesos':Procesos.objects.values('id', 'proceso').all().filter(estado=1),
            'director_list':rows_director,
            'colaborador_list': rows_colaborador,
            'lider_norma_list': rows_lider_norma
        }
        return render(request, 'admin/vincularProcesos.html',context)


def gridAreaProcesos(request):
    
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:


        #procesos = Procesos.objects.all().values().filter(estado=1, disponible=1)
        cursor = connection.cursor()
        query = " SELECT DISTINCT a.id , a.proceso,  a.tipo_proceso, "
        query = query + " CASE WHEN b.proceso_id IS NULL THEN 0 ELSE 1 END AS flag "
        query = query + " FROM areas_procesos a "
        query = query + " LEFT JOIN areas_area_proceso b ON a.id = b.proceso_id ORDER BY proceso "

        cursor.execute(query)
        rows = dictfetchall(cursor)
        
        # CONVIERTE  QuerySet  a list object
        procesos_list = list(rows)  

        if procesos_list:
            #RETORNA  
            return JsonResponse(procesos_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Procesos',
            }
            return JsonResponse(json, safe=False)
            

            
def gridVinculados(request):
    
    
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:   

        #Consultar
        cursor = connection.cursor()

        #nuevo query para adaptar a metodo vertical o transversal de procesos
        query = " SELECT  DISTINCT areas_area_proceso.id, areas_area_proceso.area_id AS id_area,  areas_area_proceso.proceso_id AS id_proceso, "
        query = query + " areas_areas.area, areas_procesos.proceso,  personal_personal.nombre AS encargado "
        query = query + " FROM areas_area_proceso "
        query = query + " LEFT JOIN areas_procesos ON areas_area_proceso.proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN areas_areas ON areas_area_proceso.area_id = areas_areas.id "
        query = query + " LEFT JOIN personal_personal ON areas_areas.id_personal_id = personal_personal.id "
        query = query + " GROUP BY areas_area_proceso.proceso_id, areas_area_proceso.area_id "
        query = query + " ORDER BY areas_areas.area "
        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        procesos_list = list(rows) 
        if procesos_list:
            #RETORNA  
            return JsonResponse(procesos_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN SECUENCIAS Y PERMITE CREAR
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Procesos Vinculados',
            }
            return JsonResponse(json, safe=False)


def insertVinculacion(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:           
        
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():
            
            #DATA RETORNADA DEL AJAX
            procesos = json.loads(request.POST.get('data'))
            for proceso in procesos['gridProcesosIngresar']:
                
                #se consulta area para clave foranea
                area=Areas.objects.get(pk=proceso['id_area'])
                
                #se selecciona el proceso para consultar si es vertical u transversal
                procesoSeleccionado=Procesos.objects.get(pk=proceso['id_proceso'])

                #si el proceso es vertical se lo establece 0->no disponible
                if procesoSeleccionado.tipo_proceso =="1":
                    estado_disponible=0
                
                #si el proceso es transversal se coloca 1->disponible
                elif procesoSeleccionado.tipo_proceso =="0":
                    estado_disponible=1 

                Procesos.objects.filter(pk=proceso['id_proceso']).update(
                    #id_area_id=area,
                    disponible=estado_disponible,
                    id_area_id=None,
                    usuario_mod_id=request.session['idUsuario'],
                    fec_modifica=proceso['datetime'],
                )
                #asignar datos al objecto proceso
                #se separa en nueva tabla por sugerencia 
                #abierta para procesos verticales y transversales
                proceso_area=Area_Proceso(
                    area_id=proceso['id_area'],
                    proceso_id=proceso['id_proceso']
                )
                #grabar proceso
                proceso_area.save()


            data = {
                'resultado':'ok_insert',
                'mensaje': 'Datos Guardados Correctamente !!',
            }
                
            return HttpResponse(json.dumps(data), content_type="application/json")

        else:
            return render(request, 'seguridad/login.html', {})


def removeVinculacion(request):
    
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:       
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():
            
            
            id_proceso = request.POST.get('id_proceso')
            id_area = request.POST.get('id_area')

            #consulta procedimiento vinculado
            #procedimiento = Area_Proceso.objects.get(area_id=id_area, proceso_id=id_proceso)
            #from areas.models import Personal
            #desvincular Personal
            #Personal.objects.filter(procedimiento_id=procedimiento.id).delete()
            #desvincular Procedimientos
            Area_Proceso.objects.filter(area_id=id_area, proceso_id=id_proceso).delete()

            data = {
                'resultado':'ok_update',
                'mensaje': 'Proceso Desvinculado !!',
            }
                
            return HttpResponse(json.dumps(data), content_type="application/json")
        
        else:
            return render(request, 'seguridad/login.html', {})
    

def excelVinculados(request):

     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        #Consultar
        cursor = connection.cursor()
        #nuevo query para adaptarse a los cambios vertical y transversal
        query = " SELECT areas_areas.area, areas_procesos.proceso,  personal_personal.nombre AS encargado, areas_procesos.estado,  "
        query = query + " personalColaborador.nombre AS colaborador "
        query = query + " FROM areas_area_proceso "
        query = query + " LEFT JOIN areas_procesos ON areas_area_proceso.proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN areas_areas ON areas_area_proceso.area_id = areas_areas.id "
        query = query + " LEFT JOIN personal_personal ON areas_areas.id_personal_id = personal_personal.id "
        query = query + " LEFT JOIN personal_personal personalColaborador ON areas_area_proceso.personal_id = personalColaborador.id "
        query = query + " GROUP BY areas_area_proceso.area_id, areas_area_proceso.proceso_id ,areas_area_proceso.personal_id"

        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        vinculados_list = list(rows)
        

        #ESTABLECER BORDES
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

        #Creamos el libro de trabajo
        wb = Workbook()

        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "Procesos Vinculados"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws.merge_cells('B2:E2')
        ws['B2'] = 'Reporte General de Procesos'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=2).border = thin_border
        b2=ws['B2']
        b2.font = Font( bold=True)
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = '#'
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=3, column=2).alignment = Alignment(horizontal='center')
        b3=ws['B3']
        b3.font = Font( bold=True)

        ws['C3'] = 'Proceso'
        ws.cell(row=3, column=3).border = thin_border
        ws.cell(row=3, column=3).alignment = Alignment(horizontal='center')
        c3=ws['C3']
        c3.font = Font( bold=True)

        ws['D3'] = 'Encargado del Proceso'
        ws.cell(row=3, column=4).border = thin_border
        ws.cell(row=3, column=4).alignment = Alignment(horizontal='center')
        d3=ws['D3']
        d3.font = Font( bold=True)

        ws['E3'] = 'Procedimiento'
        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=3, column=5).alignment = Alignment(horizontal='center')
        e3=ws['E3']
        e3.font = Font( bold=True)

        ws['F3'] = 'Colaborador'
        ws.cell(row=3, column=6).border = thin_border
        ws.cell(row=3, column=6).alignment = Alignment(horizontal='center')
        f3=ws['F3']
        f3.font = Font( bold=True)
        

        cont = 4
        indice = 1
        for vinculado in vinculados_list :

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=2).value = indice
            ws.cell(row=cont,column=2).border = thin_border
            ws.cell(row=cont,column=2).alignment = Alignment(horizontal='center')
            

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=3).value = vinculado['area']
            ws.cell(row=cont,column=3).border = thin_border
            ws.cell(row=cont,column=3).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=4).value = vinculado['encargado']
            ws.cell(row=cont,column=4).border = thin_border
            ws.cell(row=cont,column=4).alignment = Alignment(horizontal='center')

            #agregamos la data, borde, alineacion
            ws.cell(row=cont,column=5).value = vinculado['proceso']
            ws.cell(row=cont,column=5).border = thin_border
            ws.cell(row=cont,column=5).alignment = Alignment(horizontal='center')

            if vinculado['colaborador'] == None:
                ws.cell(row=cont,column=6).value = "--"
                ws.cell(row=cont,column=6).border = thin_border
                ws.cell(row=cont,column=6).alignment = Alignment(horizontal='center')
            else:
                ws.cell(row=cont,column=6).value = vinculado['colaborador']
                ws.cell(row=cont,column=6).border = thin_border
                ws.cell(row=cont,column=6).alignment = Alignment(horizontal='center')

            if vinculado['estado'] == "0":

                #obtengo celdas de estado inactivo
                areaCell=ws.cell(row=cont,column=3)
                encargadoCell=ws.cell(row=cont,column=4)
                procesoCell=ws.cell(row=cont,column=5)
                colaboradorCell=ws.cell(row=cont,column=6)


                #color rojo a inactivos
                areaCell.font = Font(color=colors.RED)
                encargadoCell.font = Font(color=colors.RED)
                procesoCell.font = Font(color=colors.RED)
                colaboradorCell.font = Font(color=colors.RED)

            cont = cont + 1
            indice = indice + 1

        #Establecemos el nombre del archivo
        nombre_archivo ="Reporte_Procesos_Vinculados.xlsx"

        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido

        #ESTABLECER DIMENSIONES A COLUMNAS
        ws.column_dimensions["C"].width = 30.0
        ws.column_dimensions["D"].width = 40.0
        ws.column_dimensions["E"].width = 40.0
        ws.column_dimensions["F"].width = 40.0

        wb.save(response)
        #retorna el archivo excel
        return response



def pdfVinculados(request):
          
     #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:
        
        return render(request, 'seguridad/login.html', {})

    else:

        #Consultar
        cursor = connection.cursor()

        #nuevo query para adaptarse a los cambios vertical y transversal
        query = " SELECT areas_areas.area, areas_procesos.proceso,  personal_personal.nombre AS encargado, areas_procesos.estado,  "
        query = query + " personalColaborador.nombre AS colaborador "
        query = query + " FROM areas_area_proceso "
        query = query + " LEFT JOIN areas_procesos ON areas_area_proceso.proceso_id = areas_procesos.id "
        query = query + " LEFT JOIN areas_areas ON areas_area_proceso.area_id = areas_areas.id "
        query = query + " LEFT JOIN personal_personal ON areas_areas.id_personal_id = personal_personal.id "
        query = query + " LEFT JOIN personal_personal personalColaborador ON areas_area_proceso.personal_id = personalColaborador.id "
        query = query + " GROUP BY areas_area_proceso.area_id, areas_area_proceso.proceso_id ,areas_area_proceso.personal_id"

        cursor.execute(query)
        rows = dictfetchall(cursor)
        # CONVIERTE  QuerySet  a list object
        vinculados_list = list(rows)  
        
        context = {
            'procesos': vinculados_list,
        }

        pdf = render_to_pdf('reportes/pdf/pdfVinculados.html', context)
        
        #FORZAR DOWNLOAD PDF
        if pdf:
            response = HttpResponse(pdf, content_type='application/pdf')
            filename = "Reporte_Vinculados_%s.pdf" %("000_000_001")
            content = "inline; filename='%s'" %(filename)

            content = "attachment; filename='%s'" %(filename)
            response['Content-Disposition'] = content
            return response


def gridColaboradores(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

            #id_cargo = request.GET['id_cargo']

            #se realiza raw sql para extrar el area vinculada
            #conexion
            
            cursor = connection.cursor()
            #query
            query = " SELECT personal_personal.id, personal_personal.nombre, personal_cargo.nombre AS cargo, areas_areas.area AS proceso, "
            query = query + " areas_procesos.proceso AS procedimiento "
            query = query + " FROM areas_area_proceso "
            query = query + " LEFT JOIN personal_personal ON areas_area_proceso.personal_id = personal_personal.id "
            query = query + " LEFT JOIN personal_cargo ON personal_personal.id_cargo_id = personal_cargo.id "
            query = query + " LEFT JOIN areas_areas ON areas_area_proceso.area_id = areas_areas.id  "
            query = query + " LEFT JOIN areas_procesos ON areas_area_proceso.proceso_id = areas_procesos.id  "
            if request.GET['id_cargo']:
                #query = query + " WHERE personal_cargo.id =%s"
                query = query + " GROUP BY nombre "
                #params=[request.GET['id_cargo']]
                cursor.execute(query)
            else:
                query = query + " GROUP BY personal_personal.nombre "
                cursor.execute(query)
            print query
            #retornando un array asociativo
            rows = dictfetchall(cursor)

            personal_list=list(rows)

            if personal_list:

                #RETORNA  
                return JsonResponse(personal_list, safe=False)

            else:
                #RETORNA MENSAJE DE NO EXISTEN
                json = {
                'resultado':'no_ok',
                'mensaje': 'no existen Colaboradores',
                }
                return JsonResponse(json, safe=False)




def selProcArea(request):
    
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

            id_area = request.POST.get('id_area')

            #procesos=Procesos.objects.values('id', 'proceso').all().filter(id_area_id=id_area, estado=1)
            
            #nuevo cambio en base a lo sugerido
            cursor = connection.cursor()
            query = " SELECT DISTINCT areas_area_proceso.proceso_id AS id, areas_procesos.proceso"
            query = query + " FROM areas_area_proceso "
            query = query + " LEFT JOIN areas_procesos  ON areas_area_proceso.proceso_id = areas_procesos.id "
            query = query + " WHERE areas_area_proceso.area_id = %s"
            #parametro id_area
            params=[id_area]
            #ejecutar consulta y retornar registros
            cursor.execute(query, params)
            rows = dictfetchall(cursor)

            procesos_list=list(rows)

            proceso = Areas.objects.values('id', 'id_personal_id').filter(pk=id_area)
            director_procesolist = list(proceso)
            if procesos_list:

                #RETORNA  
                data={
                    'resultado':'ok_select',
                    'procesos_list':procesos_list,
                    'director_list':director_procesolist
                }
                return HttpResponse(json.dumps(data), content_type="application/json")
                #return JsonResponse(procesos_list, safe=False)

            else:
                #RETORNA MENSAJE DE NO EXISTEN
                data = {
                'resultado':'no_ok',
                'mensaje': 'no existen PRocesos',
                }
                return HttpResponse(json.dumps(data), content_type="application/json")
                #return JsonResponse(json, safe=False)



def gridColabVinculados(request):
    
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        id_proceso = request.GET['id_proceso']
        id_area = request.GET['id_area']


        # query personalizado
        cursor = connection.cursor()

        query = " SELECT areas_area_proceso.id, personal_personal.nombre AS colaborador,  personal_cargo.nombre AS cargo, "
        query = query + " areas_area_proceso.rol "
        query = query + " FROM areas_area_proceso "
        query = query + " LEFT JOIN personal_personal ON areas_area_proceso.personal_id = personal_personal.id "
        query = query + " LEFT JOIN personal_cargo ON personal_personal.id_cargo_id = personal_cargo.id "
        query = query + " WHERE areas_area_proceso.area_id =%(area)s "
        query = query + " AND areas_area_proceso.proceso_id =%(proceso)s "

        params = {'area':id_area, 'proceso':id_proceso}
        cursor.execute(query, params)
        rows = dictfetchall(cursor)

        # CONVIERTE  QuerySet  a list object
        colaboradores_list = list(rows)  

        if colaboradores_list:
            #RETORNA  
            return JsonResponse(colaboradores_list, safe=False)
        else:
            #RETORNA MENSAJE DE NO EXISTEN
            json = {
            'resultado':'no_ok',
            'mensaje': 'no existen Colaboradores',
            }
            return JsonResponse(json, safe=False)


def removeColaborador(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():
                        
            
            id = request.POST.get('id')
           

            #Colaboradores.objects.filter(pk=id).delete()
            Area_Proceso.objects.filter(pk=id).update(
                personal_id=None,
                personal_disponible=1
            )
            data = {
            'resultado':'ok_delete',
            'mensaje': 'Colaborador Desvinculado !!',
            }
                
            return HttpResponse(json.dumps(data), content_type="application/json")
        
        else:

            return render(request, 'seguridad/login.html', {})
    



def insertColaboradores(request):
    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        #validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():
            import json
            dataGrid = json.loads(request.POST.get('data'))
    
            
            try:
                for dataColaborador in dataGrid['gridColaboradores']:
                    #verifica si existe colaborador vinculado a area y proceso seleccionado
                    colaborador=Area_Proceso.objects.get(personal_id=dataColaborador['id_pesonal'], proceso_id=dataColaborador['id_proceso'], area_id=dataColaborador['id_area'])
                    personal=Personal.objects.get(pk=dataColaborador['id_pesonal'])
                data = {
                    'resultado':'no_ok',
                    'mensaje': 'Colaborador Existente No pudo ser Insertado !!',
                    'personal':personal.nombre,
                }
    
                return HttpResponse(json.dumps(data), content_type="application/json")
            except Area_Proceso.DoesNotExist:


                for dataColaborador in dataGrid['gridColaboradores']:

                    existe=Area_Proceso.objects.filter(proceso_id=dataColaborador['id_proceso'], area_id=dataColaborador['id_area'], personal_id=None)

                    #si existe proceso vinculado sin personal lo actualiza
                    if existe :                       
                        #grabar
                        Area_Proceso.objects.filter(proceso_id=dataColaborador['id_proceso'], area_id=dataColaborador['id_area'], personal_id=None, personal_disponible=1).update(
                            personal_id = dataColaborador['id_pesonal'],
                            personal_disponible = 0
                        )
                        
                    else:
                        #crea un nuevo registro vinculado area/proceso/personal
                        area_proceso=Area_Proceso(
                            proceso_id=dataColaborador['id_proceso'],
                            area_id=dataColaborador['id_area'],
                            personal_id=dataColaborador['id_pesonal'],
                            personal_disponible=0
                        )
                        area_proceso.save()

                #usuario administrador
                admin = Usuario.objects.filter(id_tipo_usuario_id=1, estado=1)[0]

                for dataColaborador in dataGrid['gridColaboradores']:

                    proceso_id = dataColaborador['id_area']
                    procedimiento_id = dataColaborador['id_proceso']

                    try:
                        usuario_unificado = Usuario.objects.get(
                            director = 1,
                            colaborador = 1,
                            lider_norma = 1,
                            estado = 1,
                            id_personal_id = dataColaborador['id_pesonal']
                        )
                        if usuario_unificado:
                            #si existe usuario Director/Lider/Colaborador
                            documento = Documento(
                                    subido_por_id = usuario_unificado.id_personal_id,
                                    proceso_id = proceso_id,
                                    procedimiento_id = procedimiento_id,
                                    estado = 0
                            )
                            documento.save()

                            revision = Revision(
                                documento_id = documento.id,
                                estado_rev_director = 0,
                                estado_rev_lider = 0,
                                estado_rev_admin = 0,
                                admin_id = admin.id_personal_id,
                                director_id = usuario_unificado.id_personal_id,
                                lider_id = usuario_unificado.id_personal_id
                            )
                            revision.save()

                    except Usuario.DoesNotExist:
                        pass

                data={
                    'resultado':'ok_insert',
                    'mensaje':'Datos Grabados Exitosamente !!'
                }
                return HttpResponse(json.dumps(data), content_type="application/json")
        else:

                return render(request, 'seguridad/login.html', {})


def InsertarColaboradores(request):

    # validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        # validar si recibe los datos por ajax
        if request.method == 'POST' and request.is_ajax():

            id_area = request.POST.get('id_area')
            id_proceso = request.POST.get('id_proceso')
            director = request.POST.get('director')
            colab = request.POST.get('colaborador')
            lider_norma = request.POST.get('lider_norma')


            # verificar que no se duplique Lider de norma
            existe = Area_Proceso.objects.filter(proceso_id=id_proceso, area_id=id_area, personal_id=None)

            estado_director = 0
            estado_colaborador = 0
            estado_lidernorma = 0

            # grabar Director
            if existe:

                Area_Proceso.objects.filter(proceso_id=id_proceso, area_id=id_area, personal_id=None).update(personal_id=director, rol=1)
                estado_director = 1
            else:

                area_proceso = Area_Proceso(
                    proceso_id = id_proceso,
                    area_id = id_area,
                    personal_id = director,
                    rol = 1,
                    personal_disponible = None
                )
                area_proceso.save()
                estado_director = 1

            # grabar Colaborador
            if estado_director == 1:
                area_proceso = Area_Proceso(
                    proceso_id = id_proceso,
                    area_id = id_area,
                    personal_id = colab,
                    rol=2,
                    personal_disponible=None
                )
                area_proceso.save()
                estado_colaborador = 1

            #grabar Lider
            if estado_colaborador == 1:
                area_proceso = Area_Proceso(
                    proceso_id = id_proceso,
                    area_id = id_area,
                    personal_id = lider_norma,
                    rol=3,
                    personal_disponible=None
                )
                area_proceso.save()
                estado_lidernorma = 1

            # usuario administrador
            admin = Usuario.objects.filter(id_tipo_usuario_id=1, estado=1)[0]
            """
            from areas.models import Personal
            procedimiento = Area_Proceso.objects.get(area_id=id_area, proceso_id=id_proceso)
            director_procedimiento = Personal(
                procedimiento_id = procedimiento.id,
                personal_id = director,
                rol = 1
            )
            director_procedimiento.save()

            colaborador_procedimiento = Personal(
                procedimiento_id = procedimiento.id,
                personal_id = colab,
                rol = 2
            )
            colaborador_procedimiento.save()

            lider_norma_procedimiento = Personal(
                procedimiento_id = procedimiento.id,
                personal_id = lider_norma,
                rol = 3
            )
            lider_norma_procedimiento.save()
            """
            # si existe usuario Director/Lider/Colaborador
            documento = Documento(
                subido_por_id=colab,
                proceso_id=id_area,
                procedimiento_id=id_proceso,
                estado=0
            )
            documento.save()

            revision = Revision(
                documento_id=documento.id,
                estado_rev_director=0,
                estado_rev_lider=0,
                estado_rev_admin=0,
                admin_id=admin.id_personal_id,
                director_id=director,
                lider_id=lider_norma
            )
            revision.save()

            data = {
                'resultado': 'ok_insert',
                'mensaje': 'Datos Grabados Exitosamente !!'
            }
            return HttpResponse(json.dumps(data), content_type="application/json")
        else:

            return render(request, 'seguridad/login.html', {})




def getSubdirectorios(request):

    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        path_directorio = request.POST.get('path_directorio')
        
        import os
        from os import listdir

        mypath = path_directorio
        allfiles = listdir(mypath)
        
 

        data={
            'resultado':'ok_select',
            'subdirectorios':allfiles
        }
        return JsonResponse(data, safe=False)


def getListaArchivos(request):

    #validar si existe usuario logeado
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:

        #recibo el path/ruta absoluta
        ruta_absoluta = request.POST.get('ruta_absoluta')

        #obtengo los archivos del path
        list_dir=[]
        for dirName, subdirList, fileList in os.walk(ruta_absoluta):
            element={}
            #validamos que no vaya la carpeta root
            if dirName != 'media/gestionDocumental/':
                #nombre original
                #element['carpeta'] = dirName.split(rootDir)
                #url/ruta
                #element['url_carpeta'] = dirName
                
                #archivos
                element['archivos'] = fileList
                #agregamos element al array
                list_dir.append(element)

       

        data={
            'resultado':'ok_select',
            'archivos_list':list_dir
        }
        return JsonResponse(data, safe=False)



def reemplazarArchivo (request):
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        if request.method=='POST' and request.FILES['archivo']:
            
            form = DocumentacionForm(request.POST, request.FILES)
            myfile = request.FILES['archivo']
            #grabar archivo nuevo
            fs = FileSystemStorage(location='media/gestionDocumental/'+request.POST.get('carpeta_anterior'))
            filename = fs.save(myfile.name, myfile)
            uploaded_file_url = request.POST.get('carpeta_anterior')+"/"+filename

            nombre_carpeta = request.POST.get('carpeta_anterior')
            nombre_archivo = request.POST.get('archivo_anterior')
            ruta_absolutaAnteriorArchivo = str('media/gestionDocumental/'+nombre_carpeta+'/'+nombre_archivo)
            ruta_absolutaDestinoArchivo = str('media/gestionDocumental/'+nombre_carpeta+'/'+'Versiones'+'/'+nombre_archivo) 
            
            from shutil import Error
            # mover archivo anterior  la  la  carpeta VERSIONES
            try:

                shutil.move(ruta_absolutaAnteriorArchivo,ruta_absolutaDestinoArchivo)
            except Error as err :
                pass
            
            
            Procesos.objects.filter(pk=request.POST.get('procedimiento')).update(
                docfile=uploaded_file_url,
                estado_colaborador = 1,
                estado_encargado = 1,
                estado_lider = 1,
                observaciones_encargado = None,
                observaciones_lider = None
                #estado=0,
            )
            #borrar alerta
            if 'notificacionDocumentosPendientes' in  request.session :

                del request.session['notificacionDocumentosPendientes']
                
            return redirect('/colaborador')
        else:
            HttpResponse('ERROR SIN ARCHIVOS')

            # consultar procesos y procedimientos




def consultarProcesoPro(request):

    # validar usuario
    if 'nombreUsuario' not in request.session:

        return render(request, 'seguridad/login.html', {})

    else:
        id_pesonal = request.POST.get('id_pesonal')

        # procesos=Procesos.objects.values('id', 'proceso').all().filter(id_area_id=id_area, estado=1)

        # nuevo cambio en base a lo sugerido
        cursor = connection.cursor()
        query = " SELECT DISTINCT areas_areas.area, areas_procesos.proceso "
        query = query + " FROM areas_area_proceso "
        query = query + " LEFT JOIN personal_personal ON personal_personal.id = areas_area_proceso.personal_id "
        query = query + " LEFT JOIN areas_areas ON areas_areas.id = areas_area_proceso.area_id "
        query = query + " LEFT JOIN areas_procesos ON areas_procesos.id = areas_area_proceso.proceso_id "
        query = query + " WHERE personal_personal.id =%s"
        # parametro id_area
        params = [id_pesonal]
        # ejecutar consulta y retornar registros
        cursor.execute(query, params)
        rows_proceso = dictfetchall(cursor)

        procesosPRocedimientos = list(rows_proceso)

        if procesosPRocedimientos:
            # RETORNA
            json = {
                'resultado': 'ok_select',
                'data': procesosPRocedimientos
            }
        else:
            # RETORNA MENSAJE DE NO EXISTEN
            json = {
                'resultado': 'no_ok',
                'mensaje': 'No existen Procesos Vinculados !!',
            }
        return JsonResponse(json, safe=False)




